import argparse
import json
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

BUSINESS_PERMIT_DIR = Path(__file__).resolve().parents[1] / "BUSINESS_PERMIT_REPORT"


def clean(value):
    if value is None:
        return ""
    return str(value).strip()


def scalar(value):
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if value is None:
        return None
    return value


def decimal_value(value):
    if value in (None, ""):
        return Decimal("0")
    try:
        return Decimal(str(value).replace(",", "").strip())
    except Exception:
        return Decimal("0")


def excel_date(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return clean(value)


def latest(pattern):
    matches = sorted(BUSINESS_PERMIT_DIR.glob(pattern))
    if not matches:
        raise FileNotFoundError(f"Business permit workbook was not found: {BUSINESS_PERMIT_DIR / pattern}")
    return matches[-1]


def load_records(path, header_row=1):
    from openpyxl import load_workbook
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    headers = [clean(value) for value in next(sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True))]
    records = []
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        record = {headers[index]: row[index] for index in range(min(len(headers), len(row))) if headers[index]}
        if any(value not in (None, "") for value in record.values()):
            records.append(record)
    workbook.close()
    return records


def choose_best(existing, candidate):
    if existing is None:
        return candidate
    existing_score = (
        1 if clean(existing.get("Permit No.")) else 0,
        1 if decimal_value(existing.get("Total Amount Paid")) > 0 else 0,
        clean(existing.get("Date Issued")),
    )
    candidate_score = (
        1 if clean(candidate.get("Permit No.")) else 0,
        1 if decimal_value(candidate.get("Total Amount Paid")) > 0 else 0,
        clean(candidate.get("Date Issued")),
    )
    return candidate if candidate_score >= existing_score else existing


def build_payload(limit):
    establishment_path = latest("BUSINESS_ESTABLISHMENT-BPLS*.xlsx")
    application_path = latest("TYPE_OF_APPLICATION-BPLS*.xlsx")
    abstract_path = latest("ABSTRACT_OF_GENERAL_COLLECTION-BPLS*.xlsx")

    establishments = load_records(establishment_path, 1)
    applications = load_records(application_path, 1)
    collections = load_records(abstract_path, 7)

    application_by_id = {}
    for row in applications:
        business_id = clean(row.get("Business Identification Number"))
        if not business_id:
            continue
        current = application_by_id.get(business_id)
        paid = decimal_value(row.get("Total Amount Paid"))
        score = (1 if clean(row.get("Status of Application")) == "ISSUED" else 0, paid)
        if current is None or score >= current[0]:
            application_by_id[business_id] = (score, row)

    collection_by_id = {}
    for row in collections:
        business_id = clean(row.get("Business Identification Number"))
        if not business_id:
            continue
        bucket = collection_by_id.setdefault(business_id, {"amount_paid": Decimal("0"), "business_tax": Decimal("0"), "or_numbers": [], "latest_or_date": None})
        bucket["amount_paid"] += decimal_value(row.get("Amount Paid"))
        bucket["business_tax"] += decimal_value(row.get("Business Tax"))
        or_number = clean(row.get("O.R. Number"))
        if or_number and or_number not in bucket["or_numbers"]:
            bucket["or_numbers"].append(or_number)
        or_date = excel_date(row.get("O.R. Date"))
        if or_date and (bucket["latest_or_date"] is None or str(or_date) > str(bucket["latest_or_date"])):
            bucket["latest_or_date"] = or_date

    best_establishments = {}
    for row in establishments:
        business_id = clean(row.get("Business Identification Number"))
        if business_id:
            best_establishments[business_id] = choose_best(best_establishments.get(business_id), row)

    records = []
    for business_id, row in best_establishments.items():
        app = application_by_id.get(business_id, (None, {}))[1]
        coll = collection_by_id.get(business_id, {})
        amount = decimal_value(row.get("Total Amount Paid")) or decimal_value(app.get("Total Amount Paid")) or coll.get("amount_paid", Decimal("0"))
        status = clean(app.get("Status of Application")) or clean(row.get("Status")) or ("ISSUED" if amount > 0 else "PENDING")
        records.append({
            "business_id": business_id,
            "permit_no": clean(row.get("Permit No.")),
            "business_name": clean(row.get("Business Name")),
            "owner_name": clean(app.get("Name of Owner")) or " ".join(part for part in [clean(row.get("First Name")), clean(row.get("Middle Name")), clean(row.get("Last Name")), clean(row.get("Extension Name"))] if part),
            "barangay": clean(row.get("Barangay (Business Address")) or clean(row.get("Barangay (Business Address)")) or clean(row.get("Location of Business")).split(",")[0],
            "location": clean(row.get("Location of Business")),
            "application_date": excel_date(row.get("Application Date")),
            "application_type": clean(row.get("Type of Application")) or clean(app.get("Type of Application")),
            "business_type": clean(row.get("Type of Business")) or clean(app.get("Type of Business")),
            "business_nature": clean(row.get("Business Nature")),
            "business_line": clean(row.get("Business Line")),
            "tax_year": clean(row.get("Tax Year")),
            "or_number": clean(row.get("OR Number")) or " / ".join(coll.get("or_numbers", [])[:3]),
            "or_date": excel_date(row.get("OR Date")) or coll.get("latest_or_date"),
            "amount_paid": amount,
            "business_tax": coll.get("business_tax", Decimal("0")),
            "capital_investment": decimal_value(row.get("Capital Investment")) or decimal_value(app.get("Capital Investment")),
            "gross_sales": decimal_value(row.get("Gross Sales")) or decimal_value(app.get("Gross Sales Essential")) + decimal_value(app.get("Gross Sales Non-Essential")),
            "status": status,
            "source_type": clean(row.get("Source Type")) or clean(app.get("Source Type")),
        })

    records.sort(key=lambda item: (item.get("application_date") or "", item.get("business_name") or ""), reverse=True)
    total_revenue = sum(decimal_value(item.get("amount_paid")) for item in records)
    status_counts = {}
    application_counts = {}
    barangay_counts = {}
    for item in records:
        status_counts[item["status"] or "PENDING"] = status_counts.get(item["status"] or "PENDING", 0) + 1
        application_counts[item["application_type"] or "UNKNOWN"] = application_counts.get(item["application_type"] or "UNKNOWN", 0) + 1
        barangay_counts[item["barangay"] or "UNKNOWN"] = barangay_counts.get(item["barangay"] or "UNKNOWN", 0) + 1

    return {
        "files": {
            "establishments": establishment_path.name,
            "applications": application_path.name,
            "collections": abstract_path.name,
        },
        "summary": {
            "record_count": len(records),
            "total_revenue": total_revenue,
            "status_counts": status_counts,
            "application_counts": application_counts,
            "barangay_counts": barangay_counts,
        },
        "records": records[:limit],
    }


def main():
    parser = argparse.ArgumentParser(description="Read BPLS business permit Excel report files.")
    parser.add_argument("--limit", type=int, default=1200)
    args = parser.parse_args()
    payload = {"ok": False, "source_dir": str(BUSINESS_PERMIT_DIR)}
    try:
        payload.update(build_payload(max(1, min(args.limit, 5000))))
        payload["ok"] = True
    except Exception as exc:
        payload.update({"ok": False, "error": str(exc), "error_type": exc.__class__.__name__})
    print(json.dumps(payload, default=scalar, indent=2))
    return 0 if payload["ok"] else 1


if __name__ == "__main__":
    raise SystemExit(main())
