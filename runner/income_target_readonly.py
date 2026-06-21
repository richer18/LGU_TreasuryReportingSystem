import argparse
import json
import re
import zipfile
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from xml.etree import ElementTree as ET

try:
    from openpyxl import load_workbook
except ModuleNotFoundError:
    load_workbook = None


NS = {
    "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "rel": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "pkgrel": "http://schemas.openxmlformats.org/package/2006/relationships",
}


def scalar(value):
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value


def parse_amount(value):
    if value is None:
        return None
    if isinstance(value, (int, float, Decimal)):
        return float(value)

    text = str(value).strip()
    if text == "":
        return None

    text = text.replace(",", "")
    negative = text.startswith("(") and text.endswith(")")
    text = text.strip("()")

    try:
        amount = float(text)
    except ValueError:
        return None

    return -amount if negative else amount


def row_level(raw_label):
    leading_spaces = len(raw_label) - len(raw_label.lstrip(" "))
    if leading_spaces >= 24:
        return 2
    if leading_spaces >= 8:
        return 1
    return 0


def row_kind(label, amount):
    clean = label.strip()
    if amount is None:
        return "group"
    if clean.upper().startswith("GRAND TOTAL") or clean.upper().startswith("TOTAL "):
        return "total"
    return "target"


def default_workbook_path(year):
    return Path(__file__).resolve().parents[1] / "IncomeTarget" / f"{year}_Income_Target.xlsx"


def available_workbooks(directory):
    result = []
    for path in directory.glob("*_Income_Target.xlsx"):
        year_text = path.name.split("_", 1)[0]
        if year_text.isdigit():
            result.append((int(year_text), path))
    return sorted(result)


def source_workbook_path(requested_year, explicit_path=None):
    requested_year_number = int(requested_year)
    target_path = Path(explicit_path) if explicit_path else default_workbook_path(requested_year)

    if target_path.exists():
        return target_path, requested_year_number, False

    candidates = available_workbooks(target_path.parent)
    if not candidates:
        raise FileNotFoundError(f"Income Target workbook was not found: {target_path}")

    previous_or_same = [(year, path) for year, path in candidates if year <= requested_year_number]
    if previous_or_same:
        source_year, source_path = previous_or_same[-1]
    else:
        source_year, source_path = candidates[-1]

    return source_path, source_year, source_year != requested_year_number


def projected_amount(amount, factor):
    if amount is None:
        return None
    return round(float(amount) * factor, 2)


def column_index(cell_reference):
    letters = "".join(ch for ch in cell_reference if ch.isalpha())
    result = 0
    for letter in letters:
        result = result * 26 + (ord(letter.upper()) - ord("A") + 1)
    return result


def read_shared_strings(archive):
    try:
        root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
    except KeyError:
        return []

    strings = []
    for item in root.findall("main:si", NS):
        parts = [node.text or "" for node in item.findall(".//main:t", NS)]
        strings.append("".join(parts))
    return strings


def first_sheet_path_and_title(archive):
    workbook = ET.fromstring(archive.read("xl/workbook.xml"))
    rels = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    rel_targets = {
        rel.attrib["Id"]: rel.attrib["Target"]
        for rel in rels.findall("pkgrel:Relationship", NS)
    }

    sheet = workbook.find("main:sheets/main:sheet", NS)
    if sheet is None:
        raise RuntimeError("Income Target workbook has no worksheet.")

    rel_id = sheet.attrib[f"{{{NS['rel']}}}id"]
    target = rel_targets[rel_id].lstrip("/")
    if not target.startswith("xl/"):
        target = f"xl/{target}"

    return target, sheet.attrib.get("name", "Sheet1")


def cell_value(cell, shared_strings):
    cell_type = cell.attrib.get("t", "")

    if cell_type == "inlineStr":
        return "".join(node.text or "" for node in cell.findall(".//main:t", NS))

    value_node = cell.find("main:v", NS)
    if value_node is None or value_node.text is None:
        return None

    raw_value = value_node.text
    if cell_type == "s":
        try:
            return shared_strings[int(raw_value)]
        except (ValueError, IndexError):
            return raw_value

    if cell_type in {"str", "b"}:
        return raw_value

    try:
        number = float(raw_value)
    except ValueError:
        return raw_value

    return int(number) if number.is_integer() else number


def read_rows_without_openpyxl(workbook_path):
    rows = []
    with zipfile.ZipFile(workbook_path) as archive:
        shared_strings = read_shared_strings(archive)
        sheet_path, sheet_title = first_sheet_path_and_title(archive)
        root = ET.fromstring(archive.read(sheet_path))

        for row in root.findall(".//main:sheetData/main:row", NS):
            row_number = int(row.attrib.get("r", len(rows) + 1))
            values = {}
            for cell in row.findall("main:c", NS):
                index = column_index(cell.attrib.get("r", ""))
                if index:
                    values[index] = cell_value(cell, shared_strings)
            rows.append((row_number, values))

    return sheet_title, rows


def income_target_rows(workbook_path):
    if load_workbook is None:
        return read_rows_without_openpyxl(workbook_path)

    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    worksheet = workbook.active
    rows = []
    for row_number in range(2, worksheet.max_row + 1):
        rows.append(
            (
                row_number,
                {
                    1: worksheet.cell(row_number, 1).value,
                    3: worksheet.cell(row_number, 3).value,
                },
            )
        )
    workbook.close()
    return worksheet.title, rows


def read_income_target(args):
    workbook_path, source_year, is_projection = source_workbook_path(args.year, args.path)
    requested_year = int(args.year)
    projection_years = requested_year - int(source_year)
    projection_rate = 0.10
    projection_factor = (1 + projection_rate) ** projection_years if is_projection else 1.0

    sheet_title, sheet_rows = income_target_rows(workbook_path)

    rows = []
    current_section = ""
    for row_number, row_values in sheet_rows:
        if row_number < 2:
            continue

        raw_label = scalar(row_values.get(1))
        if raw_label is None or str(raw_label).strip() == "":
            continue

        raw_label = str(raw_label)
        label = re.sub(r"\s+", " ", raw_label.strip())
        base_amount = parse_amount(row_values.get(3))
        amount = projected_amount(base_amount, projection_factor)
        kind = row_kind(label, amount)
        level = row_level(raw_label)

        if level == 0 and kind == "group":
            current_section = label

        rows.append(
            {
                "row_number": row_number,
                "particular": label,
                "target_amount": amount,
                "base_amount": base_amount,
                "level": level,
                "kind": kind,
                "section": current_section or label,
            }
        )

    summary_labels = {
        "TAX REVENUES": "tax_revenues",
        "NON-TAX REVENUES": "non_tax_revenues",
        "TOTAL INCOME-LOCAL SOURCES": "local_sources",
        "TOTAL INCOME/RECEIPTS FROM EXTERNAL SOURCES": "external_sources",
        "TOTAL GENERAL FUND": "general_fund",
        "TOTAL SPECIAL EDUCATION FUND": "special_education_fund",
        "GRAND TOTAL (GF + SEF)": "grand_total",
    }
    summary = {}
    for row in rows:
        key = summary_labels.get(row["particular"].upper())
        if key:
            summary[key] = row["target_amount"] or 0.0

    return {
        "year": args.year,
        "source_year": str(source_year),
        "workbook": str(workbook_path),
        "sheet": sheet_title,
        "reader": "openpyxl" if load_workbook is not None else "stdlib_xlsx",
        "row_count": len(rows),
        "projection": {
            "is_projected": is_projection,
            "annual_increase_rate": projection_rate,
            "years_from_source": projection_years,
            "factor": round(projection_factor, 6),
        },
        "summary": summary,
        "rows": rows,
    }


def main():
    parser = argparse.ArgumentParser(description="Read-only Income Target workbook reader.")
    parser.add_argument("--year", default="2026")
    parser.add_argument("--path")
    args = parser.parse_args()

    payload = {
        "ok": False,
        "mode": "income_target_readonly",
    }

    try:
        payload["data"] = read_income_target(args)
        payload["ok"] = True
    except Exception as exc:
        payload.update(
            {
                "ok": False,
                "error": str(exc),
                "error_type": exc.__class__.__name__,
            }
        )

    print(json.dumps(payload, indent=2))
    return 0 if payload["ok"] else 1


if __name__ == "__main__":
    raise SystemExit(main())
