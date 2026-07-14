import sys
from pathlib import Path
from openpyxl import load_workbook
root = Path(r"\\MAIN-SERVER\LGU_TreasuryReportingSystem_$")
sys.path.insert(0, str(root / "runner"))
from rcd_access_store import TEMPLATE_PATH, fill_rcd_sheet, find_label_row

batch = {
    "id": "RCD-100-2026-377",
    "gf_rcd_no": "RCD-100-2026-377",
    "sef_rcd_no": "RCD-200-2026-377",
    "date": "2026-07-13",
    "collector": "EMILY",
    "total": 28268.35,
    "form": {"depositBank": "LAND BANK"},
}
lines = [
    {"formType": "Comm Tax.", "receiptFrom": "1552058", "receiptTo": "1552062", "collectorAmount": 216.60, "beginningQty": 38, "beginningFrom": "15521063", "beginningTo": "15521100", "receiptAccountQty": 0, "issuedQty": 5, "issuedFrom": "1552058", "issuedTo": "1552062", "endingQty": 33, "endingFrom": "15521063", "endingTo": "15521100"},
    {"formType": "AF 51", "receiptFrom": "0498062", "receiptTo": "0498100", "collectorAmount": 27094.22, "beginningQty": 0, "receiptAccountQty": 50, "receiptAccountFrom": "0498051", "receiptAccountTo": "0498100", "issuedQty": 39, "issuedFrom": "0498062", "issuedTo": "0498100", "endingQty": 11, "endingFrom": "0498101", "endingTo": "0498100"},
    {"formType": "AF 56", "receiptFrom": "2201602", "receiptTo": "2201603", "collectorAmount": 957.53, "beginningQty": 47, "beginningFrom": "2201604", "beginningTo": "2201650", "receiptAccountQty": 0, "issuedQty": 2, "issuedFrom": "2201602", "issuedTo": "2201603", "endingQty": 45, "endingFrom": "2201604", "endingTo": "2201650"},
]
wb = load_workbook(TEMPLATE_PATH)
fill_rcd_sheet(wb["100_GF"], batch, lines, "100_GF")
fill_rcd_sheet(wb["200_SEF"], batch, [lines[-1]], "200_SEF")
for name in ["100_GF", "200_SEF"]:
    ws = wb[name]
    print(f"--- {name} ---")
    for r in range(10, 46):
        vals = [ws.cell(r, c).value for c in range(1, 14)]
        if any(v not in (None, "") for v in vals):
            print(r, vals)
    label_row = find_label_row(ws, "Name and Signature of Accountable Officer")
    print("signature", label_row, ws[f"A{label_row-2}"].value if label_row else None)
