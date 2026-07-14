import sys
from pathlib import Path
from openpyxl import load_workbook

root = Path(r"\\MAIN-SERVER\LGU_TreasuryReportingSystem_$")
sys.path.insert(0, str(root / "runner"))
from rcd_access_store import TEMPLATE_PATH, fill_rcd_sheet, find_label_row

batch = {
    "id": "RCD-100-2026-377",
    "gf_rcd_no": "RCD-100-2026-377",
    "sef_rcd_no": "RCD-200-2026-377",
    "date": "2026-07-13",
    "collector": "EMILY",
    "total": 56536.70,
    "form": {},
}
lines = [
    {"formType": "Comm Tax.", "receiptFrom": "1552058", "receiptTo": "1552062", "collectorAmount": 216.60, "issuedQty": 5},
    {"formType": "AF 51", "receiptFrom": "0498062", "receiptTo": "0498100", "collectorAmount": 27094.22, "issuedQty": 39},
    {"formType": "AF 56", "receiptFrom": "2201602", "receiptTo": "2201603", "collectorAmount": 957.53, "issuedQty": 2},
]
wb = load_workbook(TEMPLATE_PATH)
fill_rcd_sheet(wb["100_GF"], batch, lines, "100_GF")
fill_rcd_sheet(wb["200_SEF"], batch, [lines[-1]], "200_SEF")
for name in ["100_GF", "200_SEF"]:
    ws = wb[name]
    label_row = find_label_row(ws, "Name and Signature of Accountable Officer")
    value = ws[f"A{label_row - 2}"].value if label_row else None
    print(name, "label_row", label_row, "signature", value)
