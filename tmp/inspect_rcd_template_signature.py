from pathlib import Path
from openpyxl import load_workbook
root = Path(r"\\MAIN-SERVER\LGU_TreasuryReportingSystem_$")
path = root / "template" / "RCD_UPDATED.xlsx"
wb = load_workbook(path)
for name in ["100_GF", "200_SEF"]:
    ws = wb[name]
    label_row = None
    for row in ws.iter_rows():
        for cell in row:
            if " ".join(str(cell.value or "").split()).strip().lower() == "name and signature of accountable officer":
                label_row = cell.row
                break
        if label_row:
            break
    print(name, "label_row", label_row, "signature_value", ws[f"A{label_row-2}"].value if label_row else None)
