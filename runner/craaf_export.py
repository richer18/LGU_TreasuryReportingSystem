import argparse
import json
import shutil
import tempfile
from copy import copy
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
TEMPLATE_PATH = PROJECT_ROOT / "template" / "CONSOLIDATED REPORT OF ACCOUNTABILITY FOR ACCOUNTABLE FORMS_TEMPLATE.xlsx"


def emit(payload, code=0):
    print(json.dumps(payload, ensure_ascii=False))
    raise SystemExit(code)


def safe_filename(value):
    cleaned = "".join(ch if ch.isalnum() or ch in "-_" else "_" for ch in str(value or "CRAAF"))
    while "__" in cleaned:
        cleaned = cleaned.replace("__", "_")
    return cleaned.strip("_") or "CRAAF"


def parse_date(value):
    text = str(value or "").strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(text[:19], fmt).date()
        except ValueError:
            pass
    return text


def period_label(payload):
    explicit = str(payload.get("period_label") or "").strip()
    if explicit:
        return explicit
    raw = payload.get("date_from") or payload.get("dateFrom") or payload.get("date_to") or payload.get("dateTo")
    parsed = parse_date(raw)
    if hasattr(parsed, "strftime"):
        return f"For the month of {parsed.strftime('%B %Y')}"
    return "For the month of ________________"


def sheet_title(payload):
    raw = payload.get("date_from") or payload.get("dateFrom") or payload.get("date_to") or payload.get("dateTo")
    parsed = parse_date(raw)
    if hasattr(parsed, "strftime"):
        return parsed.strftime("%b-%Y").upper()
    return "CRAAF"


def qty(value):
    try:
        number = int(float(str(value or 0).replace(",", "")))
    except ValueError:
        number = 0
    return number if number > 0 else None


def text(value):
    value = "" if value is None else str(value).strip()
    return value


def copy_row_style(ws, source_row, target_row):
    for col in range(1, 16):
        src = ws.cell(source_row, col)
        dst = ws.cell(target_row, col)
        if src.has_style:
            dst._style = copy(src._style)
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)


def prepare_rows(ws, needed):
    needed = max(int(needed or 0), 1)
    available = max(ws.max_row - 6, 0)
    if needed > available:
        for _ in range(needed - available):
            ws.insert_rows(ws.max_row + 1)
            copy_row_style(ws, 9, ws.max_row)
    elif needed < available:
        ws.delete_rows(7 + needed, available - needed)
    for row in range(7, 7 + needed):
        copy_row_style(ws, 7 if row == 7 else max(row - 1, 7), row)
        for col in range(1, 16):
            ws.cell(row, col).value = None


def fill_workbook(path, payload):
    rows = payload.get("rows") or []
    workbook = load_workbook(path)
    ws = workbook[workbook.sheetnames[0]]
    ws.title = sheet_title(payload)[:31]

    ws["B3"] = period_label(payload)
    prepare_rows(ws, len(rows))

    serial_cols = {4, 5, 7, 8, 10, 11, 13, 14}
    for offset, row in enumerate(rows, start=7):
        values = [
            parse_date(row.get("date")),
            text(row.get("form_type")),
            qty(row.get("beginning_qty")),
            text(row.get("beginning_from")),
            text(row.get("beginning_to")),
            qty(row.get("receipt_qty")),
            text(row.get("receipt_from")),
            text(row.get("receipt_to")),
            qty(row.get("issued_qty")),
            text(row.get("issued_from")),
            text(row.get("issued_to")),
            qty(row.get("ending_qty")),
            text(row.get("ending_from")),
            text(row.get("ending_to")),
            text(row.get("collector")),
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(offset, col)
            cell.value = value
            if col == 1 and hasattr(value, "strftime"):
                cell.number_format = "yyyy-mm-dd"
            if col in serial_cols:
                cell.number_format = "@"

    last_row = max(7, 6 + max(len(rows), 1))
    ws.print_area = f"A1:O{last_row}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    workbook.save(path)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--payload-file", required=True)
    args = parser.parse_args()

    if not TEMPLATE_PATH.exists():
        emit({"ok": False, "error": "CRAAF template was not found.", "template": str(TEMPLATE_PATH)}, 1)

    with open(args.payload_file, "r", encoding="utf-8-sig") as handle:
        payload = json.load(handle)

    out_dir = Path(tempfile.gettempdir()) / "lgu_treasury_craaf"
    out_dir.mkdir(parents=True, exist_ok=True)
    filename = f"{safe_filename('CRAAF_' + str(payload.get('date_from') or payload.get('dateFrom') or 'start') + '_' + str(payload.get('date_to') or payload.get('dateTo') or 'end'))}.xlsx"
    path = out_dir / filename
    shutil.copy2(TEMPLATE_PATH, path)
    fill_workbook(path, payload)
    emit({"ok": True, "path": str(path), "filename": filename})


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        emit({"ok": False, "error": str(exc), "type": exc.__class__.__name__}, 1)
