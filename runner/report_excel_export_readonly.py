import argparse
import json
import os
import re
import shutil
import subprocess
import sys
from datetime import datetime
from decimal import Decimal
from pathlib import Path

from firebird_probe import connect
from manual_rpt_payments_access import default_db_path as manual_rpt_db_path, list_rows as list_manual_rpt_rows
from report_preview_readonly import (
    PAID_PAYMENT_SQL,
    SUMMARY_COLUMNS,
    build_report,
    classify_summary_source,
    fetch_tax_on_business_summary,
    scalar,
)

USER_PROFILE = os.environ.get("USERPROFILE") or r"C:\Users\LIFT-LAPTOP"
USER_SITE_CANDIDATES = [
    Path(USER_PROFILE) / "AppData" / "Roaming" / "Python" / f"Python{sys.version_info.major}{sys.version_info.minor}" / "site-packages",
    Path(USER_PROFILE) / "AppData" / "Roaming" / "Python" / "Python314" / "site-packages",
    Path(USER_PROFILE) / "AppData" / "Roaming" / "Python" / "Python313" / "site-packages",
    Path(USER_PROFILE) / "AppData" / "Roaming" / "Python" / "Python312" / "site-packages",
]
for USER_SITE in USER_SITE_CANDIDATES:
    if USER_SITE.exists() and str(USER_SITE) not in sys.path:
        sys.path.append(str(USER_SITE))

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ModuleNotFoundError as exc:
    print(json.dumps({
        "ok": False,
        "mode": "read_only_excel_export",
        "error": str(exc),
        "error_type": exc.__class__.__name__,
        "python_executable": sys.executable,
        "python_path": sys.path,
        "pythonhome": os.environ.get("PYTHONHOME"),
        "pythonpath": os.environ.get("PYTHONPATH"),
    }))
    raise SystemExit(1)


TEMPLATE_DIR = Path(__file__).resolve().parents[1] / "template"
TEMPLATE_MAP = {
    21: "summary_of_collection_template.xlsx",
    22: "summary_of_collection_template_no_rpt.xlsx",
    23: "summary_of_collection_template_rpt.xlsx",
}

RECEIPT_EXCEPTION_REPORTS = {35: "canceled-void", 36: "not-remitted"}
OFFICIAL_BREAKDOWN_REPORT = 37
ESRE_QUARTERLY_REPORT = 38
OFFICIAL_CATEGORY_ORDER = [
    "Tax on Business",
    "Receipts from Economic Enterprises",
    "Regulatory Fees",
    "Service/User Charges",
]
OFFICIAL_CATEGORY_SOURCES = {
    "Tax on Business": {
        "Manufacturing",
        "Distributor",
        "Retailing",
        "Banks & Other Financial Int.",
        "Other Business Tax",
    },
    "Regulatory Fees": {
        "Mayor's Permit",
        "Weights & Measures",
        "Tricycle Permit Fee",
        "Occupation Tax",
        "Cert. of Ownership",
        "Cert. of Transfer",
        "Sand & Gravel",
        "Fines & Penalties",
        "Docking and Mooring Fee",
        "Fishing Permit Fee",
        "Miscellaneous",
    },
    "Receipts from Economic Enterprises": {
        "Water Fee",
        "Water Fees",
        "Market Stall Fee",
        "Cash Tickets",
        "SlaughterHouse Fee",
        "Slaughterhouse Fee",
        "Rental of Equipment",
        "Rent of Equipment",
        "Cockpit Share",
        "Sultadas",
        "Diving Fee",
    },
    "Service/User Charges": {
        "Registration of Birth",
        "Marriage Fee",
        "Marriage Fees",
        "Burial Fee",
        "Burial Fees",
        "Correction of Entry",
        "Sale of Agri. Prod.",
        "Sale of Acct. Forms",
        "Sale of Acc. Forms",
        "Doc Stamp Tax",
        "Secretaries Fees",
        "Secretary Fees",
        "Med./Lab. Fees",
        "Garbage Fees",
    },
}
ESRE_QUARTERLY_SECTIONS = [
    (
        "Tax on Business",
        [
            ("Manufacturing", ["Manufacturing"]),
            ("Distributor", ["Distributor"]),
            ("Retailing", ["Retailing"]),
            ("Banks & Other Financial Int.", ["Banks & Other Financial Int."]),
            ("Other Business Tax", ["Other Business Tax"]),
            ("Fines & Penalties", ["Fines & Penalties"]),
            ("Sand & Gravel", ["Sand & Gravel"]),
        ],
    ),
    (
        "Regulatory Fees and Charges",
        [
            ("WEIGHS AND MEASURE", ["Weights & Measures"]),
            ("TRICYCLE PERMIT FEE", ["Tricycle Permit Fee"]),
            ("OCCUPATION TAX", ["Occupation Tax"]),
            ("OTHER PERMITS AND LICENSE", [
                "Cockpit Share",
                "Docking and Mooring Fee",
                "Fishing Permit Fee",
                "Miscellaneous",
                "Sale of Agri. Prod.",
                "Sale of Acct. Forms",
                "Sale of Acc. Forms",
                "Sultadas",
                "Diving Fee",
            ]),
            ("CIVIL REGISTRATION", [
                "Registration of Birth",
                "Marriage Fee",
                "Marriage Fees",
                "Burial Fee",
                "Burial Fees",
                "Correction of Entry",
            ]),
            ("CATTLE/ANIMAL REGISTRATION FEES", [
                "Cert. of Ownership",
                "Cert. of Transfer",
                "Livestock",
            ]),
            ("BUILDING PERMITS", ["Building Permit Fee", "Electrical Permit Fee"]),
            ("BUSINESS PERMITS", ["Mayor's Permit"]),
            ("ZONING/LOCATION PERMIT FEES", ["Zoning Fee"]),
        ],
    ),
    (
        "Receipt from Economic Enterprise",
        [
            ("SLAUGHTERHOUSE OPERATIONS", ["SlaughterHouse Fee", "Slaughterhouse Fee"]),
            ("MARKET OPERATIONS", ["Market Stall Fee", "Cash Tickets"]),
            ("WATER WORK SYSTEM OPERATIONS", ["Water Fee", "Water Fees"]),
            ("LEASE/RENTAL FACILITIES", ["Rental of Equipment", "Rent of Equipment"]),
        ],
    ),
    (
        "Service/User Charges",
        [
            ("Police Clearance", ["Police Clearance"]),
            ("Secretaries Fee", ["Secretaries Fees", "Secretary Fees", "Doc Stamp Tax"]),
            ("Garbage Fees", ["Garbage Fees"]),
            ("Med./Lab Fees", ["Med./Lab. Fees"]),
        ],
    ),
    (
        "Other Taxes",
        [
            ("Community Tax", ["Com Tax Cert.", "Community Tax"]),
        ],
    ),
]
ESRE_RPT_GROUPS = [
    (
        "Real Property Tax - LAND = Agriculture",
        [
            ("GF", "Real Property Tax - Basic/Land"),
            ("SEF", "Real Property Tax - SEF/Land"),
        ],
    ),
    (
        "Real Property Tax - BLDG = Residential",
        [
            ("GF", "Real Property Tax - Basic/Bldg."),
            ("SEF", "Real Property Tax - SEF/Bldg."),
        ],
    ),
]
ESRE_BPLS_TAX_BUSINESS_LABELS = {
    "Manufacturing",
    "Distributor",
    "Retailing",
    "Banks & Other Financial Int.",
    "Other Business Tax",
    "Fines & Penalties",
}
RECEIPT_EXCEPTION_DEFINITIONS = {
    35: {
        "title": "Canceled / Void Receipts Report",
        "headers": [
            "OR Date",
            "OR Number",
            "Taxpayer Name",
            "Amount",
            "Fund Type",
            "Transaction Type",
            "Collector / Cashier",
            "Status",
            "Status Code",
            "Void Flag",
            "Remarks",
            "Transaction Date",
            "User ID",
        ],
        "fields": [
            "or_date",
            "or_number",
            "taxpayer_name",
            "amount",
            "fund_type",
            "transaction_type",
            "collector_cashier",
            "status",
            "status_code",
            "void_flag",
            "remarks",
            "transaction_date",
            "user_id",
        ],
    },
    36: {
        "title": "Receipts Not Remitted Report",
        "headers": [
            "OR Date",
            "OR Number",
            "Taxpayer Name",
            "Amount",
            "Fund Type",
            "Transaction Type",
            "Collector / Cashier",
            "Transaction Date",
            "RCD Number",
            "RCD Date",
            "RCD Status",
            "Days Unremitted",
            "Remarks",
        ],
        "fields": [
            "or_date",
            "or_number",
            "taxpayer_name",
            "amount",
            "fund_type",
            "transaction_type",
            "collector_cashier",
            "transaction_date",
            "rcd_number",
            "rcd_date",
            "rcd_status",
            "days_unremitted",
            "remarks",
        ],
    },
}

COLLECTOR_RECEIPT_REPORT = 34
PARENT_DELEGATED_REPORTS = set(range(1, 21)) | {26, 27, 28, 31, 32, 33}
NATIVE_TEMPLATE_REPORTS = {25, 29, 30, 39}


def parent_runner_candidates():
    this_file = Path(__file__).resolve()
    lgu_root = this_file.parents[1] if len(this_file.parents) > 1 else this_file.parent
    desktop_root = this_file.parents[2] if len(this_file.parents) > 2 else lgu_root.parent
    user_home = Path(USER_PROFILE)
    candidates = []
    candidates.extend([
        desktop_root / "ESRE_REPORT" / "run_collection_query.py",
        lgu_root.parent / "ESRE_REPORT" / "run_collection_query.py",
        user_home / "OneDrive" / "Desktop" / "ESRE_REPORT" / "run_collection_query.py",
        user_home / "Desktop" / "ESRE_REPORT" / "run_collection_query.py",
        desktop_root / "run_collection_query.py",
    ])

    unique = []
    seen = set()
    for candidate in candidates:
        resolved = str(candidate)
        if resolved not in seen:
            unique.append(candidate)
            seen.add(resolved)
    return unique


def resolve_parent_collection_runner():
    candidates = parent_runner_candidates()
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return candidates[0]


PARENT_COLLECTION_RUNNER = resolve_parent_collection_runner()


def excel_value(value):
    if isinstance(value, Decimal):
        return float(value)
    return value


def safe_filename(value):
    clean = re.sub(r"[^A-Za-z0-9_.-]+", "_", value.strip())
    return clean.strip("._") or "report"


def long_date_label(value):
    try:
        parsed = datetime.strptime(str(value), "%Y-%m-%d")
        return f"{parsed.strftime('%B')} {parsed.day} {parsed.year}"
    except (TypeError, ValueError):
        return str(value or "")


def official_category_for_source(source_name):
    for category, sources in OFFICIAL_CATEGORY_SOURCES.items():
        if source_name in sources:
            return category
    return None


def period_label(date_from, date_to):
    try:
        start_date = datetime.strptime(date_from, "%Y-%m-%d")
        end_date = datetime.strptime(date_to, "%Y-%m-%d")
        start_label = f"{start_date.strftime('%B')} {start_date.day}, {start_date.year}"
        end_label = f"{end_date.strftime('%B')} {end_date.day}, {end_date.year}"
        return start_label if date_from == date_to else f"{start_label} to {end_label}"
    except ValueError:
        return f"{date_from} to {date_to}"


def quarter_label(date_from, date_to):
    try:
        start_date = datetime.strptime(date_from, "%Y-%m-%d")
        end_date = datetime.strptime(date_to, "%Y-%m-%d")
        if start_date.year == end_date.year and start_date.month in (1, 4, 7, 10):
            expected_end_month = start_date.month + 2
            if expected_end_month == 12:
                expected_end_day = (datetime(start_date.year + 1, 1, 1) - datetime.resolution).day
            else:
                expected_end_day = (datetime(start_date.year, expected_end_month + 1, 1) - datetime.resolution).day
            if end_date.month == expected_end_month and end_date.day == expected_end_day:
                quarter_number = ((start_date.month - 1) // 3) + 1
                suffix = {1: "st", 2: "nd", 3: "rd"}.get(quarter_number, "th")
                return f"{quarter_number}{suffix} Quarter {start_date.year}"
    except ValueError:
        pass
    return "Quarterly Period"


def quarter_filename_token(date_from):
    try:
        start_date = datetime.strptime(date_from, "%Y-%m-%d")
        quarter_number = ((start_date.month - 1) // 3) + 1
        suffix = {1: "ST", 2: "ND", 3: "RD"}.get(quarter_number, "TH")
        return f"{quarter_number}{suffix}"
    except ValueError:
        return "SELECTED"


def decimal_amount(value):
    if isinstance(value, Decimal):
        return value
    if value is None:
        return Decimal("0")
    return Decimal(str(value))


def summary_row_lookup(rows):
    by_source = {}
    by_section_source = {}
    current_section = None

    for row in rows:
        source = row.get("source")
        if row.get("section"):
            current_section = source
            continue
        if not source:
            continue

        by_source.setdefault(source, []).append(row)
        if current_section:
            by_section_source[(current_section, source)] = row

    return by_source, by_section_source


def source_total(by_source, source_names, amount_column="total_collections"):
    amount = Decimal("0")
    matched_sources = []

    for source_name in source_names:
        for row in by_source.get(source_name, []):
            amount += decimal_amount(row.get(amount_column))
            matched_sources.append(source_name)

    if matched_sources:
        basis = ", ".join(dict.fromkeys(matched_sources))
    else:
        basis = "No mapped Report 21 source row found"

    return amount, basis


def bpls_tax_business_lookup(rows):
    return {
        row.get("category"): {
            "business_tax": decimal_amount(row.get("business_tax")),
            "surcharge": decimal_amount(row.get("surcharge")),
            "total": decimal_amount(row.get("total")),
        }
        for row in rows
        if row.get("category")
    }


def bpls_tax_basis_rows(rows):
    headers = ["Category", "Business Tax", "Fines & Penalties / Surcharge", "Total"]
    values = [
        [
            row.get("category", ""),
            decimal_amount(row.get("business_tax")),
            decimal_amount(row.get("surcharge")),
            decimal_amount(row.get("total")),
        ]
        for row in rows
    ]
    return headers, values


def bpls_tax_business_amount(bpls_lookup, label):
    values = bpls_lookup.get(label)
    if not values:
        return Decimal("0"), "BPLS Report 33 category not found"

    if label == "Fines & Penalties":
        return values["total"], "BPLS Business Tax: Fines & Penalties / Surcharge"

    return values["total"], "BPLS Business Tax: Business Tax"


def rpt_line_amount(by_section_source, section_name, line_name, amount_column):
    row = by_section_source.get((section_name, line_name))
    if not row:
        return Decimal("0"), f"{section_name} / {line_name} not found in Report 21"

    return decimal_amount(row.get(amount_column)), f"{section_name} / {line_name} / {amount_column}"


def esre_rpt_summary_basis_rows(report_rows):
    _, by_section_source = summary_row_lookup(report_rows)
    headers = ["ESRE Group", "Fund", "Summary of Collection Section", "Line", "Amount"]
    rows = []
    for group_name, fund_groups in ESRE_RPT_GROUPS:
        for fund_name, summary_section in fund_groups:
            for display_line, summary_line in (("Current", "Current Year"), ("Prior", "Previous Years"), ("Penalties", "Penalties")):
                amount, _basis = rpt_line_amount(
                    by_section_source,
                    summary_section,
                    summary_line,
                    "total_collections",
                )
                rows.append([group_name, fund_name, summary_section, display_line, amount])
    return headers, rows


def rpt_cell_lookup(cells):
    return {
        (int(cell["row"]), int(cell["column"])): decimal_amount(cell.get("value"))
        for cell in cells
    }


def fetch_esre_rpt_collection_buckets(date_from, date_to):
    report_year = datetime.strptime(date_from, "%Y-%m-%d").year
    buckets = {}
    sql = f"""
        SELECT
            COALESCE(pcd.PROPERTYKIND_CT, prop.PROPERTYKIND_CT) AS PROPERTYKIND_CT,
            kind.DESCRIPTION AS PROPERTYKIND_NAME,
            COALESCE(pcd.CLASSCODE_CT, ra.PREDOMCLASSCODE_CT) AS CLASSCODE_CT,
            cls.DESCRIPTION AS CLASSIFICATION_NAME,
            pcd.ITAXTYPE_CT,
            pcd.CASETYPE_CT,
            pcd.TAXYEAR,
            SUM(pcd.AMOUNT) AS AMOUNT
        FROM PAYMENT p
        JOIN PAYMENTCLASSDETAIL pcd ON pcd.PAYMENT_ID = p.PAYMENT_ID
        LEFT JOIN RPTASSESSMENT ra ON ra.TAXTRANS_ID = pcd.TAXTRANS_ID
        LEFT JOIN PROPERTY prop ON prop.PROP_ID = ra.PROP_ID
        LEFT JOIN T_PROPERTYKIND kind ON kind.CODE = COALESCE(pcd.PROPERTYKIND_CT, prop.PROPERTYKIND_CT)
        LEFT JOIN T_CLASSIFICATION cls ON cls.CODE = COALESCE(pcd.CLASSCODE_CT, ra.PREDOMCLASSCODE_CT)
        WHERE p.PAYMENTDATE >= CAST(? AS DATE)
          AND p.PAYMENTDATE < DATEADD(1 DAY TO CAST(? AS DATE))
          AND p.PAYGROUP_CT = 'RPT'
          {PAID_PAYMENT_SQL}
          AND COALESCE(pcd.CANCELLED_BV, 0) = 0
          AND TRIM(pcd.ITAXTYPE_CT) IN ('BSC', 'SEF')
        GROUP BY COALESCE(pcd.PROPERTYKIND_CT, prop.PROPERTYKIND_CT),
                 kind.DESCRIPTION,
                 COALESCE(pcd.CLASSCODE_CT, ra.PREDOMCLASSCODE_CT),
                 cls.DESCRIPTION,
                 pcd.ITAXTYPE_CT, pcd.CASETYPE_CT, pcd.TAXYEAR
    """

    connection = connect()
    try:
        cursor = connection.cursor()
        cursor.execute(sql, (date_from, date_to))
        for property_kind, property_kind_name, class_code, classification_name, tax_type, case_type, taxyear, amount in cursor.fetchall():
            property_kind = (property_kind or "").strip().upper()
            class_key = (class_code or "").strip().upper()
            tax_type = (tax_type or "").strip()
            case_type = (case_type or "").strip()
            amount = amount or Decimal("0")

            key = (property_kind, class_key, tax_type)
            if key not in buckets:
                buckets[key] = {
                    "property_kind": property_kind,
                    "property_kind_name": (property_kind_name or property_kind or "Unknown").strip(),
                    "class_code": class_key,
                    "classification_name": (classification_name or class_key or "Unknown").strip(),
                    "current": Decimal("0"),
                    "prior": Decimal("0"),
                    "penalties": Decimal("0"),
                    "advance": Decimal("0"),
                }

            if case_type == "PEN":
                if taxyear and taxyear > report_year:
                    buckets[key]["advance"] += amount
                else:
                    buckets[key]["penalties"] += amount
            elif taxyear and taxyear > report_year:
                buckets[key]["advance"] += amount
            elif case_type == "DED":
                if taxyear == report_year:
                    buckets[key]["current"] -= abs(amount)
                else:
                    buckets[key]["prior"] -= abs(amount)
            elif taxyear == report_year:
                buckets[key]["current"] += amount
            else:
                buckets[key]["prior"] += amount
        connection.rollback()
    finally:
        connection.close()

    return buckets


def esre_rpt_property_label(bucket):
    property_kind = (bucket.get("property_kind") or "").upper()
    property_name = (bucket.get("property_kind_name") or property_kind or "Unknown").strip()
    labels = {
        "L": "LAND",
        "B": "BLDG",
        "M": "MACHINERY",
        "P": "IMPROVEMENTS",
    }
    return labels.get(property_kind, property_name.upper())


def esre_rpt_classification_label(bucket):
    classification = (bucket.get("classification_name") or bucket.get("class_code") or "Unknown").strip()
    return classification.title()


def esre_rpt_group_key(bucket):
    property_order = {"L": 1, "B": 2, "P": 3, "M": 4}
    class_order = {"A": 1, "R": 2, "C": 3, "I": 4, "SS": 5}
    return (
        property_order.get((bucket.get("property_kind") or "").upper(), 99),
        class_order.get((bucket.get("class_code") or "").upper(), 99),
        esre_rpt_property_label(bucket),
        esre_rpt_classification_label(bucket),
    )


def esre_rpt_grouped_buckets(buckets):
    grouped = {}
    for (property_kind, class_key, tax_type), bucket in buckets.items():
        group_key = (property_kind, class_key)
        group = grouped.setdefault(group_key, {
            "property_kind": bucket.get("property_kind"),
            "property_kind_name": bucket.get("property_kind_name"),
            "class_code": bucket.get("class_code"),
            "classification_name": bucket.get("classification_name"),
            "funds": {},
        })
        group["funds"][tax_type] = bucket
    return sorted(grouped.values(), key=esre_rpt_group_key)


def esre_rpt_collection_basis_rows(buckets):
    headers = ["Property Kind", "Property Class", "Fund", "Current", "Prior", "Penalties", "Advance Excluded"]
    rows = []
    for group in esre_rpt_grouped_buckets(buckets):
        for tax_type, fund in (("BSC", "GF"), ("SEF", "SEF")):
            bucket = group["funds"].get(tax_type, {})
            rows.append([
                esre_rpt_property_label(group),
                esre_rpt_classification_label(group),
                fund,
                bucket.get("current", Decimal("0")),
                bucket.get("prior", Decimal("0")),
                bucket.get("penalties", Decimal("0")),
                bucket.get("advance", Decimal("0")),
            ])
    return headers, rows


def rpt_class_line_amount(cell_lookup, template_row, tax_type, line_name, share_rate):
    if tax_type == "BSC":
        columns = {"current": 3, "discount": 4, "prior": 5, "penalty_current": 6, "penalty_prior": 7}
    else:
        columns = {"current": 10, "discount": 11, "prior": 12, "penalty_current": 13, "penalty_prior": 14}

    current = cell_lookup.get((template_row, columns["current"]), Decimal("0"))
    discount = cell_lookup.get((template_row, columns["discount"]), Decimal("0"))
    prior = cell_lookup.get((template_row, columns["prior"]), Decimal("0"))
    penalty_current = cell_lookup.get((template_row, columns["penalty_current"]), Decimal("0"))
    penalty_prior = cell_lookup.get((template_row, columns["penalty_prior"]), Decimal("0"))

    if line_name == "Current":
        applicable_discount = min(current, discount)
        gross_amount = current - applicable_discount
        basis_column = f"{tax_type} current less applicable discount"
    elif line_name == "Prior":
        gross_amount = prior
        basis_column = f"{tax_type} prior"
    else:
        gross_amount = penalty_current + penalty_prior
        basis_column = f"{tax_type} penalties current + prior"

    return gross_amount * share_rate, f"RPT classification row {template_row}; {basis_column}; share rate {share_rate:.0%}"


def rpt_basis_rows(cells):
    cell_lookup = rpt_cell_lookup(cells)
    definitions = [
        ("Land Agriculture", 11),
        ("Building Residential", 23),
    ]
    headers = [
        "Property Class",
        "Template Row",
        "BSC Current",
        "BSC Discount",
        "BSC Prior",
        "BSC Penalty Current",
        "BSC Penalty Prior",
        "SEF Current",
        "SEF Discount",
        "SEF Prior",
        "SEF Penalty Current",
        "SEF Penalty Prior",
    ]
    rows = []
    for label, row_index in definitions:
        rows.append([
            label,
            row_index,
            cell_lookup.get((row_index, 3), Decimal("0")),
            cell_lookup.get((row_index, 4), Decimal("0")),
            cell_lookup.get((row_index, 5), Decimal("0")),
            cell_lookup.get((row_index, 6), Decimal("0")),
            cell_lookup.get((row_index, 7), Decimal("0")),
            cell_lookup.get((row_index, 10), Decimal("0")),
            cell_lookup.get((row_index, 11), Decimal("0")),
            cell_lookup.get((row_index, 12), Decimal("0")),
            cell_lookup.get((row_index, 13), Decimal("0")),
            cell_lookup.get((row_index, 14), Decimal("0")),
        ])
    return headers, rows


def build_esre_quarterly_rows(report_rows, bpls_tax_rows, rpt_collection_buckets):
    by_source, by_section_source = summary_row_lookup(report_rows)
    bpls_lookup = bpls_tax_business_lookup(bpls_tax_rows)
    rows = []
    grand_total = Decimal("0")

    for section_name, lines in ESRE_QUARTERLY_SECTIONS:
        line_rows = []
        section_total = Decimal("0")
        for label, source_names in lines:
            if section_name == "Tax on Business" and label in ESRE_BPLS_TAX_BUSINESS_LABELS:
                amount, basis = bpls_tax_business_amount(bpls_lookup, label)
            else:
                amount, basis = source_total(by_source, source_names)
            section_total += amount
            line_rows.append({
                "category": "",
                "particular": label,
                "amount": amount,
                "basis": basis,
                "level": 1,
            })

        grand_total += section_total
        rows.append({
            "category": section_name,
            "particular": "",
            "amount": section_total,
            "basis": "Subtotal of rows below",
            "level": 0,
        })
        rows.extend(line_rows)

    rpt_line_names = (("Current", "Current Year"), ("Prior", "Previous Years"), ("Penalties", "Penalties"))
    for group_name, fund_groups in ESRE_RPT_GROUPS:
        group_rows = []
        group_total = Decimal("0")
        for fund_name, summary_section in fund_groups:
            fund_rows = []
            fund_total = Decimal("0")
            for display_line, summary_line in rpt_line_names:
                amount, basis = rpt_line_amount(
                    by_section_source,
                    summary_section,
                    summary_line,
                    "total_collections",
                )
                fund_total += amount
                fund_rows.append({
                    "category": "",
                    "particular": f"{display_line}",
                    "amount": amount,
                    "basis": basis,
                    "level": 2,
                })

            group_total += fund_total
            group_rows.append({
                "category": "",
                "particular": fund_name,
                "amount": fund_total,
                "basis": f"Subtotal of Summary of Collection section: {summary_section}",
                "level": 1,
            })
            group_rows.extend(fund_rows)

        grand_total += group_total
        rows.append({
            "category": group_name,
            "particular": "",
            "amount": group_total,
            "basis": "Subtotal of GF and SEF rows below",
            "level": 0,
        })
        rows.extend(group_rows)

    report_21_total = Decimal("0")
    for report_row in report_rows:
        if report_row.get("source") == "TOTAL" and not report_row.get("section"):
            report_21_total = decimal_amount(report_row.get("total_collections"))
            break

    reconciliation_amount = report_21_total - grand_total
    if abs(reconciliation_amount) > Decimal("0.004"):
        rows.append({
            "category": "Reconciliation Difference",
            "particular": "",
            "amount": reconciliation_amount,
            "basis": "Report 21 TOTAL less displayed ESRE category subtotals",
            "level": 0,
        })
        grand_total += reconciliation_amount

    rows.append({
        "category": "Grand Total",
        "particular": "",
        "amount": grand_total,
        "basis": "Matches Report 21 TOTAL row",
        "level": 0,
        "is_total": True,
    })

    return rows, grand_total


def adjusted_rpt_amount(case_type, amount):
    amount = amount or Decimal("0")
    if (case_type or "").strip() == "DED":
        return -abs(amount)
    return amount


def official_breakdown_data(date_from, date_to):
    category_totals = {category: Decimal("0") for category in OFFICIAL_CATEGORY_ORDER}
    detail_rows = []

    non_rpt_sql = f"""
        SELECT
            CAST(p.PAYMENTDATE AS DATE) AS OR_DATE,
            TRIM(p.RECEIPTNO) AS OR_NUMBER,
            TRIM(p.PAIDBY) AS TAXPAYER_NAME,
            COALESCE(NULLIF(TRIM(p.COLLECTOR), ''), TRIM(p.USERID), 'UNSPECIFIED') AS COLLECTOR_CASHIER,
            p.PAYMENTDATE AS TRANSACTION_DATE,
            TRIM(p.PAYGROUP_CT) AS FUND_TYPE,
            TRIM(pd.ITAXTYPE_CT) AS SOURCE_CODE,
            pd.SOURCEID AS SOURCE_ID,
            TRIM(pd.SOURCE_CT) AS SOURCE_CT,
            COALESCE(TRIM(it.DESCRIPTION), TRIM(pd.ITAXTYPE_CT), '') AS SOURCE_DESCRIPTION,
            TRIM(opr.DESCRIPTION) AS CHILD_DESCRIPTION,
            pd.AMOUNTPAID AS AMOUNT
        FROM PAYMENT p
        JOIN PAYMENTDETAIL pd ON pd.PAYMENT_ID = p.PAYMENT_ID
        LEFT JOIN T_ITAXTYPE it ON it.CODE = pd.ITAXTYPE_CT
        LEFT JOIN T_OTHERPAYMENTRATE opr ON opr.OPRATE_ID = pd.SOURCEID
        WHERE p.PAYMENTDATE >= CAST(? AS DATE)
          AND p.PAYMENTDATE < DATEADD(1 DAY TO CAST(? AS DATE))
          {PAID_PAYMENT_SQL}
          AND COALESCE(TRIM(p.PAYGROUP_CT), '') <> 'RPT'
        ORDER BY CAST(p.PAYMENTDATE AS DATE), TRIM(p.RECEIPTNO), p.PAYMENT_ID, pd.RECEIPTITEMORDER
    """
    rpt_sql = f"""
        SELECT
            CAST(p.PAYMENTDATE AS DATE) AS OR_DATE,
            TRIM(p.RECEIPTNO) AS OR_NUMBER,
            TRIM(p.PAIDBY) AS TAXPAYER_NAME,
            COALESCE(NULLIF(TRIM(p.COLLECTOR), ''), TRIM(p.USERID), 'UNSPECIFIED') AS COLLECTOR_CASHIER,
            p.PAYMENTDATE AS TRANSACTION_DATE,
            COALESCE(NULLIF(TRIM(p.PAYGROUP_CT), ''), 'RPT') AS FUND_TYPE,
            TRIM(pcd.ITAXTYPE_CT) AS TAX_TYPE,
            TRIM(pcd.CASETYPE_CT) AS CASE_TYPE,
            SUM(pcd.AMOUNT) AS AMOUNT
        FROM PAYMENT p
        JOIN PAYMENTCLASSDETAIL pcd ON pcd.PAYMENT_ID = p.PAYMENT_ID
        WHERE p.PAYMENTDATE >= CAST(? AS DATE)
          AND p.PAYMENTDATE < DATEADD(1 DAY TO CAST(? AS DATE))
          AND COALESCE(TRIM(p.PAYGROUP_CT), '') = 'RPT'
          {PAID_PAYMENT_SQL}
          AND COALESCE(pcd.CANCELLED_BV, 0) = 0
          AND TRIM(pcd.ITAXTYPE_CT) IN ('BSC', 'SEF')
        GROUP BY CAST(p.PAYMENTDATE AS DATE), TRIM(p.RECEIPTNO), TRIM(p.PAIDBY),
                 COALESCE(NULLIF(TRIM(p.COLLECTOR), ''), TRIM(p.USERID), 'UNSPECIFIED'),
                 p.PAYMENTDATE, COALESCE(NULLIF(TRIM(p.PAYGROUP_CT), ''), 'RPT'),
                 TRIM(pcd.ITAXTYPE_CT), TRIM(pcd.CASETYPE_CT)
        ORDER BY CAST(p.PAYMENTDATE AS DATE), TRIM(p.RECEIPTNO)
    """

    rpt_gf = Decimal("0")
    rpt_sf = Decimal("0")
    connection = connect()
    try:
        cursor = connection.cursor()
        cursor.execute(non_rpt_sql, (date_from, date_to))
        for row in cursor.fetchall():
            (
                or_date,
                or_number,
                taxpayer_name,
                collector_cashier,
                transaction_date,
                fund_type,
                source_code,
                source_id,
                source_ct,
                source_description,
                child_description,
                amount,
            ) = row
            source_name = classify_summary_source(source_code, source_id, source_ct)
            category = official_category_for_source(source_name)
            if not category:
                continue

            amount = amount or Decimal("0")
            category_totals[category] += amount
            detail_rows.append({
                "or_date": or_date,
                "or_number": or_number,
                "taxpayer_name": taxpayer_name,
                "category": category,
                "source": source_name,
                "amount": amount,
                "fund_type": fund_type or "General Fund",
                "collector_cashier": collector_cashier,
                "transaction_date": transaction_date,
                "remarks": child_description or source_description or source_name,
            })

        cursor.execute(rpt_sql, (date_from, date_to))
        for row in cursor.fetchall():
            (
                or_date,
                or_number,
                taxpayer_name,
                collector_cashier,
                transaction_date,
                fund_type,
                tax_type,
                case_type,
                amount,
            ) = row
            net_amount = adjusted_rpt_amount(case_type, amount)
            if tax_type == "BSC":
                share = net_amount * Decimal("0.40")
                rpt_gf += share
                category = "RPT GF - Municipal Basic Share 40%"
            elif tax_type == "SEF":
                share = net_amount * Decimal("0.50")
                rpt_sf += share
                category = "RPT SF - Municipal SEF Share 50%"
            else:
                continue

            detail_rows.append({
                "or_date": or_date,
                "or_number": or_number,
                "taxpayer_name": taxpayer_name,
                "category": category,
                "source": tax_type,
                "amount": share,
                "fund_type": fund_type or "RPT",
                "collector_cashier": collector_cashier,
                "transaction_date": transaction_date,
                "remarks": f"Municipal share from {tax_type}; DED lines subtracted before share",
            })
        connection.rollback()
    finally:
        connection.close()

    rpt_municipal = rpt_gf + rpt_sf
    grand_total = sum(category_totals.values(), Decimal("0")) + rpt_municipal

    summary_rows = [
        {
            "category": category,
            "amount": category_totals[category],
            "remarks": "Paid non-RPT PAYMENTDETAIL lines grouped by existing source category mapping",
            "is_breakdown": False,
        }
        for category in OFFICIAL_CATEGORY_ORDER
    ]
    summary_rows.extend([
        {
            "category": "Real Property Tax only Municipal Sharing",
            "amount": rpt_municipal,
            "remarks": "Subtotal only: RPT GF municipal share + RPT SF municipal share",
            "is_breakdown": False,
        },
        {
            "category": "RPT GF - Municipal Basic Share 40%",
            "amount": rpt_gf,
            "remarks": "Breakdown only; not added separately to Grand Total",
            "is_breakdown": True,
        },
        {
            "category": "RPT SF - Municipal SEF Share 50%",
            "amount": rpt_sf,
            "remarks": "Breakdown only; not added separately to Grand Total",
            "is_breakdown": True,
        },
        {
            "category": "Grand Total",
            "amount": grand_total,
            "remarks": "Main categories 1-4 + Real Property Tax only Municipal Sharing",
            "is_total": True,
        },
    ])

    return summary_rows, detail_rows, {
        "rpt_gf": rpt_gf,
        "rpt_sf": rpt_sf,
        "rpt_municipal": rpt_municipal,
        "grand_total": grand_total,
    }


def write_official_breakdown_workbook(date_from, date_to, output_dir):
    summary_rows, detail_rows, totals = official_breakdown_data(date_from, date_to)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Summary"

    title_fill = PatternFill("solid", fgColor="0554F2")
    header_fill = PatternFill("solid", fgColor="EAF2FF")
    total_fill = PatternFill("solid", fgColor="EAF7EA")
    breakdown_fill = PatternFill("solid", fgColor="F8FAFC")

    sheet.merge_cells("A1:D1")
    sheet["A1"] = "OFFICIAL REPORT BREAKDOWN"
    sheet["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    sheet["A1"].fill = title_fill
    sheet["A1"].alignment = Alignment(horizontal="center")
    sheet.merge_cells("A2:D2")
    sheet["A2"] = "CATEGORY BREAKDOWN"
    sheet["A2"].font = Font(bold=True, size=12)
    sheet["A2"].alignment = Alignment(horizontal="center")
    sheet.merge_cells("A3:D3")
    sheet["A3"] = f"Period: {period_label(date_from, date_to)}"
    sheet["A3"].alignment = Alignment(horizontal="center")

    headers = ["Category", "Amount", "Percentage of Total", "Remarks / Basis"]
    header_row = 5
    for column_index, header in enumerate(headers, start=1):
        cell = sheet.cell(header_row, column_index)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    grand_total = totals["grand_total"]
    for row_index, row in enumerate(summary_rows, start=header_row + 1):
        sheet.cell(row_index, 1).value = row["category"]
        sheet.cell(row_index, 2).value = excel_value(row["amount"])
        sheet.cell(row_index, 3).value = float((row["amount"] / grand_total) if grand_total else 0)
        sheet.cell(row_index, 4).value = row["remarks"]
        sheet.cell(row_index, 2).number_format = "#,##0.00"
        sheet.cell(row_index, 3).number_format = "0.00%"

        if row.get("is_breakdown"):
            sheet.cell(row_index, 1).value = "  " + row["category"]
            for column_index in range(1, 5):
                sheet.cell(row_index, column_index).fill = breakdown_fill
                sheet.cell(row_index, column_index).font = Font(italic=True)
        if row.get("is_total"):
            for column_index in range(1, 5):
                sheet.cell(row_index, column_index).fill = total_fill
                sheet.cell(row_index, column_index).font = Font(bold=True)

    detail_sheet = workbook.create_sheet("Details")
    detail_headers = [
        "OR Date",
        "OR Number",
        "Taxpayer Name",
        "Category",
        "Source / Revenue Code",
        "Amount",
        "Fund Type",
        "Collector / Cashier",
        "Transaction Date",
        "Remarks",
    ]
    for column_index, header in enumerate(detail_headers, start=1):
        cell = detail_sheet.cell(1, column_index)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_index, row in enumerate(detail_rows, start=2):
        detail_sheet.cell(row_index, 1).value = excel_value(row["or_date"])
        detail_sheet.cell(row_index, 2).value = row["or_number"]
        detail_sheet.cell(row_index, 3).value = row["taxpayer_name"]
        detail_sheet.cell(row_index, 4).value = row["category"]
        detail_sheet.cell(row_index, 5).value = row["source"]
        detail_sheet.cell(row_index, 6).value = excel_value(row["amount"])
        detail_sheet.cell(row_index, 7).value = row["fund_type"]
        detail_sheet.cell(row_index, 8).value = row["collector_cashier"]
        detail_sheet.cell(row_index, 9).value = excel_value(row["transaction_date"])
        detail_sheet.cell(row_index, 10).value = row["remarks"]
        detail_sheet.cell(row_index, 6).number_format = "#,##0.00"

    for target_sheet in (sheet, detail_sheet):
        for column_index, column in enumerate(target_sheet.columns, start=1):
            max_length = max(len(str(cell.value or "")) for cell in column)
            target_sheet.column_dimensions[get_column_letter(column_index)].width = min(max(max_length + 2, 12), 52)

    sheet.freeze_panes = "A6"
    detail_sheet.freeze_panes = "A2"

    output_dir.mkdir(parents=True, exist_ok=True)
    output_name = safe_filename(f"report_37_official_report_breakdown_{date_from}_to_{date_to}.xlsx")
    output_path = output_dir / output_name
    workbook.save(output_path)

    return output_path, len(detail_rows), totals


def write_esre_quarterly_workbook(date_from, date_to, output_dir):
    payload = build_report(21, date_from, date_to)
    report_rows = payload["rows"]
    bpls_tax_rows = fetch_tax_on_business_summary(date_from, date_to)
    esre_rows, grand_total = build_esre_quarterly_rows(report_rows, bpls_tax_rows, {})

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "ESRE Quarterly"

    title_fill = PatternFill("solid", fgColor="1F4E79")
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    section_fill = PatternFill("solid", fgColor="EAF2F8")
    subheader_fill = PatternFill("solid", fgColor="F3F6FA")
    total_fill = PatternFill("solid", fgColor="E2F0D9")

    sheet.merge_cells("A1:C1")
    sheet["A1"] = "ESRE QUARTERLY REPORT"
    sheet["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    sheet["A1"].fill = title_fill
    sheet["A1"].alignment = Alignment(horizontal="center")
    sheet.merge_cells("A2:C2")
    sheet["A2"] = quarter_label(date_from, date_to)
    sheet["A2"].font = Font(bold=True, size=12)
    sheet["A2"].alignment = Alignment(horizontal="center")
    sheet.merge_cells("A3:C3")
    sheet["A3"] = f"Period: {period_label(date_from, date_to)}"
    sheet["A3"].alignment = Alignment(horizontal="center")

    headers = ["Category", "Particular", "Amount"]
    header_row = 5
    for column_index, header in enumerate(headers, start=1):
        cell = sheet.cell(header_row, column_index)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_index, row in enumerate(esre_rows, start=header_row + 1):
        sheet.cell(row_index, 1).value = row["category"]
        sheet.cell(row_index, 2).value = row["particular"]
        sheet.cell(row_index, 3).value = excel_value(row["amount"])
        sheet.cell(row_index, 3).number_format = "#,##0.00"

        if row.get("is_total"):
            for column_index in range(1, 4):
                sheet.cell(row_index, column_index).fill = total_fill
                sheet.cell(row_index, column_index).font = Font(bold=True)
        elif row["level"] == 0:
            for column_index in range(1, 4):
                sheet.cell(row_index, column_index).fill = section_fill
                sheet.cell(row_index, column_index).font = Font(bold=True)
        elif row["level"] == 1:
            sheet.cell(row_index, 2).value = f"  {row['particular']}"
            if row["particular"] in ("GF", "SEF"):
                for column_index in range(1, 4):
                    sheet.cell(row_index, column_index).fill = subheader_fill
                    sheet.cell(row_index, column_index).font = Font(bold=True)
        elif row["level"] == 2:
            sheet.cell(row_index, 2).value = f"    {row['particular']}"

    basis_sheet = workbook.create_sheet("Report 21 Basis")
    for column_index, header in enumerate(SUMMARY_COLUMNS, start=1):
        cell = basis_sheet.cell(1, column_index)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_index, row in enumerate(report_rows, start=2):
        row_values = summary_excel_row(row)
        for column_index, value in enumerate(row_values, start=1):
            basis_sheet.cell(row_index, column_index).value = excel_value(value)
            if column_index > 1:
                basis_sheet.cell(row_index, column_index).number_format = "#,##0.00"
        if row.get("section"):
            for column_index in range(1, len(SUMMARY_COLUMNS) + 1):
                basis_sheet.cell(row_index, column_index).fill = subheader_fill
                basis_sheet.cell(row_index, column_index).font = Font(bold=True)

    bpls_basis_sheet = workbook.create_sheet("BPLS Tax Basis")
    bpls_headers, bpls_rows = bpls_tax_basis_rows(bpls_tax_rows)
    for column_index, header in enumerate(bpls_headers, start=1):
        cell = bpls_basis_sheet.cell(1, column_index)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for row_index, row_values in enumerate(bpls_rows, start=2):
        for column_index, value in enumerate(row_values, start=1):
            bpls_basis_sheet.cell(row_index, column_index).value = excel_value(value)
            if column_index > 1:
                bpls_basis_sheet.cell(row_index, column_index).number_format = "#,##0.00"

    rpt_collection_sheet = workbook.create_sheet("RPT Summary Basis")
    rpt_headers, rpt_rows = esre_rpt_summary_basis_rows(report_rows)
    for column_index, header in enumerate(rpt_headers, start=1):
        cell = rpt_collection_sheet.cell(1, column_index)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for row_index, row_values in enumerate(rpt_rows, start=2):
        for column_index, value in enumerate(row_values, start=1):
            rpt_collection_sheet.cell(row_index, column_index).value = excel_value(value)
            if column_index > 4:
                rpt_collection_sheet.cell(row_index, column_index).number_format = "#,##0.00"

    sheet.delete_cols(4)

    for target_sheet in (sheet, basis_sheet, bpls_basis_sheet, rpt_collection_sheet):
        for column_index, column in enumerate(target_sheet.columns, start=1):
            max_length = max(len(str(cell.value or "")) for cell in column)
            target_sheet.column_dimensions[get_column_letter(column_index)].width = min(max(max_length + 2, 12), 56)

    sheet.freeze_panes = "A6"
    basis_sheet.freeze_panes = "A2"
    bpls_basis_sheet.freeze_panes = "A2"
    rpt_collection_sheet.freeze_panes = "A2"

    output_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    output_name = safe_filename(f"ESRE-REPORT-{quarter_filename_token(date_from)}-QTR-{timestamp}.xlsx")
    output_path = output_dir / output_name
    workbook.save(output_path)

    return output_path, len(esre_rows), grand_total


def summary_excel_row(row):
    if row.get("section"):
        return [row.get("source", "")] + [""] * (len(SUMMARY_COLUMNS) - 1)

    return [row.get(column, "") for column in SUMMARY_COLUMNS]


def write_summary_workbook(report_number, payload, date_from, date_to, output_dir):
    template_name = TEMPLATE_MAP[report_number]
    template_path = TEMPLATE_DIR / template_name

    if not template_path.exists():
        raise FileNotFoundError(f"Template file was not found: {template_path}")

    workbook = load_workbook(template_path)
    sheet = workbook.active

    try:
        start_date = datetime.strptime(date_from, "%Y-%m-%d")
        end_date = datetime.strptime(date_to, "%Y-%m-%d")
        month_label = f"Month of {start_date.strftime('%B %Y')}"
        if start_date.month != end_date.month or start_date.year != end_date.year:
            month_label = f"{start_date.strftime('%B %d, %Y')} to {end_date.strftime('%B %d, %Y')}"
        sheet["A4"] = month_label
    except ValueError:
        sheet["A4"] = f"{date_from} to {date_to}"

    start_row = 8
    body_rows = [summary_excel_row(row) for row in payload["rows"]]

    for index, row_values in enumerate(body_rows):
        excel_row = start_row + index
        for col_index, value in enumerate(row_values[:12], start=1):
            sheet.cell(excel_row, col_index).value = excel_value(value)

    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True

    output_dir.mkdir(parents=True, exist_ok=True)
    output_name = safe_filename(f"report_{report_number}_{date_from}_to_{date_to}.xlsx")
    output_path = output_dir / output_name
    workbook.save(output_path)

    return output_path, len(body_rows)


def receipt_exception_args(report_number, date_from, date_to):
    class Args:
        pass

    args = Args()
    args.report = RECEIPT_EXCEPTION_REPORTS[report_number]
    args.date_from = date_from
    args.date_to = date_to
    args.fund_type = ""
    args.collector = ""
    args.status = ""
    args.transaction_type = ""
    args.or_number = ""
    args.taxpayer = ""
    args.page = 1
    args.limit = 500
    return args


def receipt_exception_rows(report_number, date_from, date_to):
    from receipt_exceptions_readonly import canceled_void_report, not_remitted_report

    args = receipt_exception_args(report_number, date_from, date_to)
    if report_number == 35:
        rows, warnings = canceled_void_report(args)
    else:
        rows, warnings = not_remitted_report(args)
    return rows, warnings


def write_receipt_exception_workbook(report_number, date_from, date_to, output_dir):
    definition = RECEIPT_EXCEPTION_DEFINITIONS[report_number]
    rows, warnings = receipt_exception_rows(report_number, date_from, date_to)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = safe_filename(definition["title"])[:31]

    sheet["A1"] = definition["title"]
    sheet["A2"] = "Municipality of Zamboanguita"
    sheet["A3"] = f"Period: {date_from} to {date_to}"
    if warnings:
        sheet["A4"] = "Warnings: " + " | ".join(warnings)

    header_row = 6
    for column_index, header in enumerate(definition["headers"], start=1):
        cell = sheet.cell(header_row, column_index)
        cell.value = header
        cell.font = cell.font.copy(bold=True)

    for row_index, row in enumerate(rows, start=header_row + 1):
        for column_index, field in enumerate(definition["fields"], start=1):
            sheet.cell(row_index, column_index).value = excel_value(row.get(field, ""))

    amount_column = definition["fields"].index("amount") + 1
    for row_index in range(header_row + 1, header_row + len(rows) + 1):
        sheet.cell(row_index, amount_column).number_format = '#,##0.00'

    for column in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        sheet.column_dimensions[column[0].column_letter].width = min(max(max_length + 2, 12), 45)

    output_dir.mkdir(parents=True, exist_ok=True)
    output_name = safe_filename(f"report_{report_number}_{date_from}_to_{date_to}_{definition['title']}.xlsx")
    output_path = output_dir / output_name
    workbook.save(output_path)

    return output_path, len(rows)


def collector_receipt_args(date_from, date_to, collector):
    class Args:
        pass

    args = Args()
    args.date_from = date_from
    args.date_to = date_to
    args.collector = collector or ""
    args.receipt_from = ""
    args.receipt_to = ""
    args.fund_scope = "all"
    args.limit = 100000
    return args


def collector_receipt_rows(date_from, date_to, collector):
    from general_fund_readonly import receipt_report

    connection = connect()
    try:
        cursor = connection.cursor()
        return receipt_report(cursor, collector_receipt_args(date_from, date_to, collector))
    finally:
        connection.close()


def write_collector_receipt_workbook(date_from, date_to, output_dir, collector=None):
    rows = collector_receipt_rows(date_from, date_to, collector)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Collector Receipts"

    title = "Generate Collection Receipt Per Collector"
    sheet.merge_cells("A1:J1")
    sheet["A1"] = title
    sheet["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    sheet["A1"].fill = PatternFill("solid", fgColor="0554F2")
    sheet["A1"].alignment = Alignment(horizontal="center")

    sheet.merge_cells("A2:J2")
    sheet["A2"] = f"Period: from {long_date_label(date_from)} to {long_date_label(date_to)}"
    sheet["A2"].alignment = Alignment(horizontal="center")

    if collector:
        sheet.merge_cells("A3:J3")
        sheet["A3"] = f"Collector: {collector}"
        sheet["A3"].alignment = Alignment(horizontal="center")

    headers = ["Date", "Collector", "Fund / Source", "Receipt Type", "OR No.", "Taxpayer", "Lines", "Status", "RCD No.", "Total"]
    header_row = 5
    header_fill = PatternFill("solid", fgColor="EAF2FF")
    for column_index, header in enumerate(headers, start=1):
        cell = sheet.cell(header_row, column_index)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    total_amount = Decimal("0")
    for row_index, row in enumerate(rows, start=header_row + 1):
        amount = Decimal(str(row.get("total_amount") or 0))
        total_amount += amount
        values = [
            row.get("collection_date"),
            row.get("collector"),
            row.get("fund_source"),
            row.get("receipt_type"),
            row.get("receipt_no"),
            row.get("taxpayer"),
            row.get("line_count"),
            row.get("collection_status"),
            row.get("rcd_number"),
            amount,
        ]
        for column_index, value in enumerate(values, start=1):
            sheet.cell(row_index, column_index).value = excel_value(value)
        sheet.cell(row_index, 10).number_format = "#,##0.00"

    total_row = header_row + len(rows) + 1
    sheet.cell(total_row, 9).value = "Total"
    sheet.cell(total_row, 9).font = Font(bold=True)
    sheet.cell(total_row, 10).value = float(total_amount)
    sheet.cell(total_row, 10).font = Font(bold=True)
    sheet.cell(total_row, 10).number_format = "#,##0.00"

    for column_index, column in enumerate(sheet.columns, start=1):
        max_length = max(len(str(cell.value or "")) for cell in column)
        sheet.column_dimensions[get_column_letter(column_index)].width = min(max(max_length + 2, 12), 52)

    sheet.freeze_panes = "A6"
    output_dir.mkdir(parents=True, exist_ok=True)
    collector_token = safe_filename(collector or "all_collectors")
    output_name = safe_filename(f"report_34_collector_receipts_{collector_token}_{date_from}_to_{date_to}.xlsx")
    output_path = output_dir / output_name
    workbook.save(output_path)

    return output_path, len(rows), total_amount


def save_workbook_with_fallback(workbook, output_path):
    try:
        workbook.save(output_path)
    except PermissionError:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = output_path.with_name(f"{output_path.stem}_{timestamp}{output_path.suffix}")
        workbook.save(output_path)
    return output_path




def write_rpt_record_workbook(rows, output_path, date_from, date_to):
    workbook = load_workbook(TEMPLATE_DIR / "RECORD OF REAL PROPERTY TAX COLLECTION.xlsx")
    sheet = workbook.active

    try:
        start_label = datetime.strptime(date_from, "%Y-%m-%d").strftime("%B %d, %Y")
        end_label = datetime.strptime(date_to, "%Y-%m-%d").strftime("%B %d, %Y")
        sheet["E4"] = start_label if date_from == date_to else f"{start_label} to {end_label}"
    except ValueError:
        sheet["E4"] = f"{date_from} to {date_to}"

    start_row = 11
    body_rows = rows[1:]
    for index, row_values in enumerate(body_rows):
        excel_row = start_row + index
        for col_index, value in enumerate(row_values[:36], start=1):
            sheet.cell(excel_row, col_index).value = excel_value(value)

    output_path = output_path.with_suffix(".xlsx")
    output_path = save_workbook_with_fallback(workbook, output_path)
    return len(body_rows), output_path




def write_abstract_general_collections_workbook(rows, daily_rows, output_path, date_from, date_to):
    workbook = load_workbook(TEMPLATE_DIR / "ABSTRACT_OF_GENERAL_COLLECTIONS.xlsx")
    data_sheet = workbook["data"]
    daily_sheet = workbook["daily_collection"]
    data_sheet["K4"] = period_label(date_from, date_to)

    for index, row_values in enumerate(rows[1:], start=8):
        for col_index, value in enumerate(row_values, start=1):
            data_sheet.cell(index, col_index).value = excel_value(value)

    for index, row_values in enumerate(daily_rows[1:], start=5):
        for col_index, value in enumerate(row_values, start=1):
            daily_sheet.cell(index, col_index).value = excel_value(value)

    output_path = output_path.with_suffix(".xlsx")
    output_path = save_workbook_with_fallback(workbook, output_path)
    return len(rows) - 1 if rows else 0, output_path


def write_abstract_trust_funds_workbook(rows, daily_rows, output_path, date_from, date_to):
    workbook = load_workbook(TEMPLATE_DIR / "ABSTRACT_OF_TRUST_FUNDS_COLLECTIONS.xlsx")
    data_sheet = workbook["data"]
    daily_sheet = workbook["daily_collection"]
    data_sheet["H4"] = period_label(date_from, date_to)

    for index, row_values in enumerate(rows[1:], start=9):
        for col_index, value in enumerate(row_values, start=1):
            data_sheet.cell(index, col_index).value = excel_value(value)

    for index, row_values in enumerate(daily_rows[1:], start=5):
        for col_index, value in enumerate(row_values, start=1):
            daily_sheet.cell(index, col_index).value = excel_value(value)

    output_path = output_path.with_suffix(".xlsx")
    output_path = save_workbook_with_fallback(workbook, output_path)
    return len(rows) - 1 if rows else 0, output_path




GENERAL_ABSTRACT_COLUMNS = {
    "Manufacturing": 4,
    "Distributor": 5,
    "Retailing": 6,
    "Banks & Other Financial Int.": 7,
    "Other Business Tax": 8,
    "Sand & Gravel": 9,
    "Fines & Penalties": 10,
    "Mayor's Permit": 11,
    "Weights & Measures": 12,
    "Tricycle Permit Fee": 13,
    "Occupation Tax": 14,
    "Cert. of Ownership": 15,
    "Cert. of Transfer": 16,
    "Docking and Mooring Fee": 19,
    "Sultadas": 20,
    "Miscellaneous": 21,
    "Registration of Birth": 22,
    "Marriage Fees": 23,
    "Burial Fees": 24,
    "Correction of Entry": 25,
    "Fishing Permit Fee": 26,
    "Sale of Agri. Prod.": 27,
    "Sale of Acc. Forms": 28,
    "Water Fees": 29,
    "Market Stall Fee": 30,
    "Slaughterhouse Fee": 32,
    "Rent of Equipment": 33,
    "Doc Stamp Tax": 34,
    "Police Clearance": 35,
    "Secretary Fees": 36,
    "Med./Lab. Fees": 37,
    "Garbage Fees": 38,
}


TRUST_ABSTRACT_NAMES = {
    "Building Permit Fee",
    "Electrical Permit Fee",
    "Zoning Fee",
    "Livestock",
    "Diving Fee",
}




def rpt_record_headers():
    return [
        "DATE",
        "PAID_BY",
        "NAME_OF_TAXPAYER",
        "PERIOD_COVERED",
        "PIN",
        "OR_NO",
        "TD_ARP_NO",
        "BARANGAY",
        "BASIC_CURRENT_YEAR_GROSS",
        "BASIC_DISCOUNT",
        "BASIC_PRIOR_YEARS",
        "BASIC_PENALTY_CURRENT_YEAR",
        "BASIC_PENALTY_PREV_YEARS",
        "BASIC_PENALTY_PRIOR_YEARS",
        "BASIC_GROSS",
        "BASIC_NET",
        "SEF_CURRENT_YEAR_GROSS",
        "SEF_DISCOUNT",
        "SEF_PRIOR_YEARS",
        "SEF_PENALTY_CURRENT_YEAR",
        "SEF_PENALTY_PREV_YEARS",
        "SEF_PENALTY_PRIOR_YEARS",
        "SEF_GROSS",
        "SEF_NET",
        "GRAND_GROSS",
        "GRAND_NET",
        "BASIC_25_PERCENT_SHARE",
        "PROPERTY_CLASSIFICATION",
        "PROPERTY_KIND",
        "COLLECTOR",
        "PAYMENT_STATUS_CT",
        "IS_CANCELLED",
        "PAYMENT_TOTAL_AMOUNT",
        "BOOKINGREFERENCE",
        "IS_VOID",
        "INCLUDE_IN_REPORT",
    ]


def period_covered(taxyears):
    years = sorted(year for year in taxyears if year is not None)
    if not years:
        return ""
    if len(years) == 1:
        return str(years[0])
    if years == list(range(years[0], years[-1] + 1)):
        return f"{years[0]}-{years[-1]}"
    return ", ".join(str(year) for year in years)


def display_lookup(description, code):
    description = (description or "").strip()
    code = (code or "").strip()
    if description:
        return description
    return code


def property_classification_label(description, code):
    code = (code or "").strip()
    if code.upper().startswith("S"):
        return "SPECIAL"
    return display_lookup(description, code)


def property_classification_key(code):
    code = (code or "").strip().upper()
    if code.startswith("S"):
        return "SPECIAL"
    return code


def add_rpt_record_amount(record, tax_type, case_type, taxyear, amount, current_taxyear):
    tax_prefix = "basic" if tax_type == "BSC" else "sef" if tax_type == "SEF" else None
    if tax_prefix is None:
        return

    amount = amount or 0
    if case_type == "DED":
        record[f"{tax_prefix}_discount"] += abs(amount)
    elif case_type == "PEN":
        if taxyear == current_taxyear:
            record[f"{tax_prefix}_pen_current"] += amount
        elif current_taxyear is not None and taxyear == current_taxyear - 1:
            record[f"{tax_prefix}_pen_prev"] += amount
        else:
            record[f"{tax_prefix}_pen_prior"] += amount
    else:
        if taxyear == current_taxyear:
            record[f"{tax_prefix}_current"] += amount
        else:
            record[f"{tax_prefix}_prior"] += amount


def manual_rpt_record_rows(date_from, date_to, current_taxyear):
    try:
        manual_rows = list_manual_rpt_rows(
            manual_rpt_db_path(),
            date_from=date_from,
            date_to=date_to,
            limit=10000,
        )
    except Exception:
        return []

    rows = []
    for row in manual_rows:
        tax_year = row.get("taxyear")
        try:
            tax_year_int = int(tax_year) if tax_year not in (None, "") else None
        except (TypeError, ValueError):
            tax_year_int = None
        basic_tax = Decimal(str(row.get("basic_tax") or 0))
        basic_penalty = Decimal(str(row.get("basic_penalty") or 0))
        sef_tax = Decimal(str(row.get("sef_tax") or 0))
        sef_penalty = Decimal(str(row.get("sef_penalty") or 0))
        basic_current = basic_tax if tax_year_int == current_taxyear else Decimal("0")
        basic_prior = Decimal("0") if tax_year_int == current_taxyear else basic_tax
        basic_pen_current = basic_penalty if tax_year_int == current_taxyear else Decimal("0")
        basic_pen_prior = Decimal("0") if tax_year_int == current_taxyear else basic_penalty
        sef_current = sef_tax if tax_year_int == current_taxyear else Decimal("0")
        sef_prior = Decimal("0") if tax_year_int == current_taxyear else sef_tax
        sef_pen_current = sef_penalty if tax_year_int == current_taxyear else Decimal("0")
        sef_pen_prior = Decimal("0") if tax_year_int == current_taxyear else sef_penalty
        basic_gross = basic_current + basic_prior + basic_pen_current + basic_pen_prior
        sef_gross = sef_current + sef_prior + sef_pen_current + sef_pen_prior
        grand_gross = basic_gross + sef_gross
        rows.append([
            row.get("payment_date"),
            row.get("paid_by"),
            row.get("declared_owner") or row.get("paid_by"),
            str(tax_year or ""),
            "",
            row.get("receipt_no"),
            row.get("td_no"),
            "",
            basic_current,
            Decimal("0"),
            basic_prior,
            basic_pen_current,
            Decimal("0"),
            basic_pen_prior,
            basic_gross,
            basic_gross,
            sef_current,
            Decimal("0"),
            sef_prior,
            sef_pen_current,
            Decimal("0"),
            sef_pen_prior,
            sef_gross,
            sef_gross,
            grand_gross,
            Decimal(str(row.get("total_amount") or grand_gross)),
            basic_gross * Decimal("0.25"),
            "Manual RPT",
            "Manual RPT Payment",
            row.get("collector"),
            "MANUAL",
            0,
            Decimal(str(row.get("total_amount") or grand_gross)),
            row.get("rcd_number"),
            0,
            1,
        ])
    return rows


def build_rpt_record_rows_from_fdb(date_from, date_to, user, password):
    current_taxyear = datetime.strptime(date_from, "%Y-%m-%d").year
    sql = """
        SELECT
            p.PAYMENT_ID,
            p.PAYMENTDATE,
            p.PAIDBY,
            tx.OWNERNAME,
            p.RECEIPTNO,
            p.STATUS_CT,
            p.AMOUNT,
            p.RCDNUMBER,
            p.VOID_BV,
            COALESCE(p.COLLECTOR, p.USERID) AS COLLECTOR_NAME,
            pcd.TAXTRANS_ID,
            pcd.ITAXTYPE_CT,
            pcd.CASETYPE_CT,
            pcd.TAXYEAR,
            pcd.AMOUNT AS DETAIL_AMOUNT,
            pcd.CANCELLED_BV,
            pcd.CLASSCODE_CT,
            pcd.PROPERTYKIND_CT,
            ra.TDNO,
            ra.TDNOFORGR,
            ra.PREDOMCLASSCODE_CT,
            prop.PINNO,
            prop.NEWPINNO,
            brgy.DESCRIPTION AS BARANGAY_NAME,
            cls.DESCRIPTION AS CLASSIFICATION_NAME,
            kind.DESCRIPTION AS PROPERTY_KIND_NAME
        FROM PAYMENT p
        JOIN PAYMENTCLASSDETAIL pcd ON pcd.PAYMENT_ID = p.PAYMENT_ID
        LEFT JOIN TAXPAYER tx ON tx.LOCAL_TIN = p.LOCAL_TIN
        LEFT JOIN RPTASSESSMENT ra ON ra.TAXTRANS_ID = pcd.TAXTRANS_ID
        LEFT JOIN PROPERTY prop ON prop.PROP_ID = ra.PROP_ID
        LEFT JOIN T_BARANGAY brgy
               ON brgy.CODE = prop.BARANGAY_CT
              AND brgy.MUNICIPAL_ID = prop.MUNICIPAL_ID
              AND brgy.PROVINCE_CT = prop.PROVINCE_CT
        LEFT JOIN T_CLASSIFICATION cls
               ON cls.CODE = COALESCE(pcd.CLASSCODE_CT, ra.PREDOMCLASSCODE_CT)
        LEFT JOIN T_PROPERTYKIND kind
               ON kind.CODE = COALESCE(pcd.PROPERTYKIND_CT, prop.PROPERTYKIND_CT)
        WHERE p.PAYMENTDATE >= CAST(? AS DATE)
          AND p.PAYMENTDATE < DATEADD(1 DAY TO CAST(? AS DATE))
          AND p.PAYGROUP_CT = 'RPT'
          AND COALESCE(p.VOID_BV, 0) = 0
          AND COALESCE(TRIM(p.STATUS_CT), '') NOT IN ('CNL', 'CAN', 'CNC', 'CANCEL', 'CANCELLED', 'VOID', 'VOI')
          AND COALESCE(pcd.CANCELLED_BV, 0) = 0
        ORDER BY p.PAYMENTDATE, p.RECEIPTNO, p.PAYMENT_ID, pcd.TAXTRANS_ID, pcd.TAXYEAR
    """
    amount_fields = (
        "basic_current", "basic_discount", "basic_prior",
        "basic_pen_current", "basic_pen_prev", "basic_pen_prior",
        "sef_current", "sef_discount", "sef_prior",
        "sef_pen_current", "sef_pen_prev", "sef_pen_prior",
    )
    records = {}
    order = []

    con = connect()
    try:
        cur = con.cursor()
        cur.execute(sql, (date_from, date_to))
        for (
            payment_id, payment_date, paid_by, taxpayer_name, receipt_no, status_ct,
            payment_amount, booking_reference, void_bv, collector_name, taxtrans_id,
            tax_type, case_type, taxyear, detail_amount, cancelled_bv, class_code,
            property_kind_code, td_no, td_no_for_gr, predom_class_code, pin_no,
            new_pin_no, barangay_name, classification_name, property_kind_name,
        ) in cur.fetchall():
            classification_code = class_code or predom_class_code
            key = (payment_id, taxtrans_id, property_classification_key(classification_code))
            if key not in records:
                records[key] = {
                    "payment_date": payment_date,
                    "paid_by": paid_by,
                    "taxpayer_name": taxpayer_name or paid_by,
                    "taxyears": set(),
                    "pin": new_pin_no or pin_no,
                    "receipt_no": receipt_no,
                    "td_arp_no": td_no_for_gr or td_no,
                    "barangay": barangay_name,
                    "classification": property_classification_label(
                        classification_name,
                        classification_code,
                    ),
                    "property_kind": display_lookup(property_kind_name, property_kind_code),
                    "collector": collector_name,
                    "status_ct": status_ct,
                    "is_cancelled": cancelled_bv or 0,
                    "payment_amount": payment_amount or 0,
                    "booking_reference": booking_reference,
                    "is_void": void_bv or 0,
                    "include_in_report": 1,
                }
                for field in amount_fields:
                    records[key][field] = Decimal("0")
                order.append(key)

            record = records[key]
            record["taxyears"].add(taxyear)
            if not record["classification"]:
                record["classification"] = property_classification_label(
                    classification_name,
                    classification_code,
                )
            if not record["property_kind"]:
                record["property_kind"] = display_lookup(property_kind_name, property_kind_code)
            add_rpt_record_amount(
                record,
                (tax_type or "").strip(),
                (case_type or "").strip(),
                taxyear,
                detail_amount,
                current_taxyear,
            )
        con.rollback()
    finally:
        con.close()

    rows = [rpt_record_headers()]
    for key in order:
        record = records[key]
        basic_gross = (
            record["basic_current"] + record["basic_prior"] +
            record["basic_pen_current"] + record["basic_pen_prev"] + record["basic_pen_prior"]
        )
        basic_net = basic_gross - record["basic_discount"]
        sef_gross = (
            record["sef_current"] + record["sef_prior"] +
            record["sef_pen_current"] + record["sef_pen_prev"] + record["sef_pen_prior"]
        )
        sef_net = sef_gross - record["sef_discount"]
        grand_gross = basic_gross + sef_gross
        grand_net = basic_net + sef_net
        rows.append([
            record["payment_date"],
            record["paid_by"],
            record["taxpayer_name"],
            period_covered(record["taxyears"]),
            record["pin"],
            record["receipt_no"],
            record["td_arp_no"],
            record["barangay"],
            record["basic_current"],
            record["basic_discount"],
            record["basic_prior"],
            record["basic_pen_current"],
            record["basic_pen_prev"],
            record["basic_pen_prior"],
            basic_gross,
            basic_net,
            record["sef_current"],
            record["sef_discount"],
            record["sef_prior"],
            record["sef_pen_current"],
            record["sef_pen_prev"],
            record["sef_pen_prior"],
            sef_gross,
            sef_net,
            grand_gross,
            grand_net,
            basic_net * Decimal("0.25"),
            record["classification"],
            record["property_kind"],
            record["collector"],
            record["status_ct"],
            record["is_cancelled"],
            record["payment_amount"],
            record["booking_reference"],
            record["is_void"],
            record["include_in_report"],
        ])
    rows.extend(manual_rpt_record_rows(date_from, date_to, current_taxyear))
    return rows




def payment_detail_rows_for_abstract(date_from, date_to, user, password):
    sql = """
        SELECT
            p.PAYMENT_ID,
            p.PAYMENTDATE,
            p.RECEIPTNO,
            p.PAIDBY,
            COALESCE(p.COLLECTOR, p.USERID) AS COLLECTOR_NAME,
            p.PAYGROUP_CT,
            pd.ITAXTYPE_CT,
            pd.SOURCEID,
            pd.SOURCE_CT,
            pd.AMOUNTPAID
        FROM PAYMENT p
        JOIN PAYMENTDETAIL pd ON pd.PAYMENT_ID = p.PAYMENT_ID
        WHERE p.PAYMENTDATE >= CAST(? AS DATE)
          AND p.PAYMENTDATE < DATEADD(1 DAY TO CAST(? AS DATE))
          AND COALESCE(p.VOID_BV, 0) = 0
          AND COALESCE(TRIM(p.STATUS_CT), '') NOT IN ('CNL', 'CAN', 'CNC', 'CANCEL', 'CANCELLED', 'VOID', 'VOI')
          AND COALESCE(p.PAYGROUP_CT, '') <> 'RPT'
        ORDER BY p.PAYMENTDATE, p.RECEIPTNO, p.PAYMENT_ID, pd.RECEIPTITEMORDER
    """
    rows = []
    con = connect()
    try:
        cur = con.cursor()
        cur.execute(sql, (date_from, date_to))
        for row in cur.fetchall():
            rows.append(row)
        con.rollback()
    finally:
        con.close()
    return rows


def add_general_abstract_amount(record, source_name, amount):
    amount = amount or 0
    if source_name == "Cockpit Share":
        record[17] += amount * Decimal("0.50")
        record[18] += amount * Decimal("0.50")
        return True

    col_index = GENERAL_ABSTRACT_COLUMNS.get(source_name)
    if col_index is None:
        return False
    record[col_index] += amount
    return True


def build_abstract_general_collections_rows_from_fdb(date_from, date_to, user, password):
    records = {}
    order = []

    for (
        payment_id, payment_date, receipt_no, paid_by, collector_name, paygroup,
        itaxtype, source_id, source_ct, amount,
    ) in payment_detail_rows_for_abstract(date_from, date_to, user, password):
        source_name = classify_summary_source(
            itaxtype.strip() if isinstance(itaxtype, str) else itaxtype,
            source_id,
            source_ct.strip() if isinstance(source_ct, str) else source_ct,
        )
        if source_name in TRUST_ABSTRACT_NAMES or source_name == "Community Tax":
            continue

        if payment_id not in records:
            records[payment_id] = {
                "date": payment_date,
                "receipt_no": receipt_no,
                "paid_by": paid_by,
                "collector": collector_name,
                "paygroup": paygroup,
                "amounts": {col_index: Decimal("0") for col_index in range(4, 39)},
            }
            order.append(payment_id)

        added = add_general_abstract_amount(records[payment_id]["amounts"], source_name, amount)
        if not added and source_name:
            records[payment_id]["amounts"][21] += amount or 0

    headers = [
        "Date", "Receipt Number", "Names", "Manufacturing", "Distributor", "Retailing",
        "Financial", "Other", "Sand & Gravel", "Fines & Penalties", "Mayor's Permit",
        "W. & M.", "Trirycle Operators", "Occu.", "Cert. of Ownership", "Cert. of Transfer",
        "Cockpit Prov. Share", "Cockpit Local Share", "Docking and Mooring Fee", "Sultadas",
        "MISCS.", "Reg. of", "Marriage Fees", "Burial Fees", "Correction of Entry",
        "Fishing Permit Fee", "Sale of Agri. Prod.", "Sale of Acct. Form", "Water Fees",
        "Stall Fees", "Cash Tickets", "Slaughter House Fee", "Rental of Equipment",
        "Doc. Stamp", "Police Clearance", "Cert.", "Med./Dent. & Lab. Fees", "Garbage Fees",
        "CASHIER", "TOTAL", "TYPE OF RECIEPT",
    ]
    rows = [headers]
    daily = {}

    for payment_id in order:
        record = records[payment_id]
        total = sum(record["amounts"].values())
        if total == 0:
            continue
        row = [
            record["date"], record["receipt_no"], record["paid_by"],
            *[record["amounts"].get(col_index, Decimal("0")) for col_index in range(4, 39)],
            record["collector"], total, record["paygroup"],
        ]
        rows.append(row)

        day = record["date"].date() if hasattr(record["date"], "date") else record["date"]
        if day not in daily:
            daily[day] = {col_index: Decimal("0") for col_index in range(4, 39)}
        for col_index, value in record["amounts"].items():
            daily[day][col_index] += value

    daily_rows = [["Date"] + headers[3:38] + ["TOTAL"]]
    for day in sorted(daily):
        total = sum(daily[day].values())
        daily_rows.append([day] + [daily[day].get(col_index, Decimal("0")) for col_index in range(4, 39)] + [total])
    return rows, daily_rows


def trust_split_values(source_name, amount):
    amount = amount or 0
    values = {col_index: Decimal("0") for col_index in range(4, 14)}
    if source_name == "Building Permit Fee":
        values[4] = amount * Decimal("0.80")
        values[5] = amount * Decimal("0.15")
        values[6] = amount * Decimal("0.05")
    elif source_name == "Electrical Permit Fee":
        values[7] = amount
    elif source_name == "Zoning Fee":
        values[8] = amount
    elif source_name == "Livestock":
        values[9] = amount * Decimal("0.80")
        values[10] = amount * Decimal("0.20")
    elif source_name == "Diving Fee":
        values[11] = amount * Decimal("0.40")
        values[12] = amount * Decimal("0.30")
        values[13] = amount * Decimal("0.30")
    return values


def build_abstract_trust_funds_rows_from_fdb(date_from, date_to, user, password):
    records = {}
    order = []

    for (
        payment_id, payment_date, receipt_no, paid_by, collector_name, paygroup,
        itaxtype, source_id, source_ct, amount,
    ) in payment_detail_rows_for_abstract(date_from, date_to, user, password):
        source_name = classify_summary_source(
            itaxtype.strip() if isinstance(itaxtype, str) else itaxtype,
            source_id,
            source_ct.strip() if isinstance(source_ct, str) else source_ct,
        )
        if source_name not in TRUST_ABSTRACT_NAMES:
            continue

        if payment_id not in records:
            records[payment_id] = {
                "date": payment_date,
                "receipt_no": receipt_no,
                "paid_by": paid_by,
                "collector": collector_name,
                "paygroup": paygroup,
                "amounts": {col_index: Decimal("0") for col_index in range(4, 14)},
            }
            order.append(payment_id)

        split_values = trust_split_values(source_name, amount)
        for col_index, value in split_values.items():
            records[payment_id]["amounts"][col_index] += value

    headers = [
        "Date", "Receipt Number", "Names", "Building Fee 80% Local",
        "Building Fee 15% T.F.", "Building Fee 5% Nat'L.", "Electrical Fee",
        "Zoning Fee", "Livestock 80% Local", "Livestock 20% Nat'l",
        "Diving 40% GF", "Diving 30% Fishers", "Diving 30% Brgy",
        "CASHIER", "Total", "TYPE OF RECIEPT",
    ]
    rows = [headers]
    daily = {}

    for payment_id in order:
        record = records[payment_id]
        total = sum(record["amounts"].values())
        if total == 0:
            continue
        row = [
            record["date"], record["receipt_no"], record["paid_by"],
            *[record["amounts"].get(col_index, Decimal("0")) for col_index in range(4, 14)],
            record["collector"], total, record["paygroup"],
        ]
        rows.append(row)

        day = record["date"].date() if hasattr(record["date"], "date") else record["date"]
        if day not in daily:
            daily[day] = {col_index: Decimal("0") for col_index in range(4, 14)}
        for col_index, value in record["amounts"].items():
            daily[day][col_index] += value

    daily_rows = [["DATE"] + headers[3:13] + ["Total"]]
    for day in sorted(daily):
        total = sum(daily[day].values())
        daily_rows.append([day] + [daily[day].get(col_index, Decimal("0")) for col_index in range(4, 14)] + [total])
    return rows, daily_rows




def database_rows(sql, params):
    connection = connect()
    try:
        cursor = connection.cursor()
        cursor.execute(sql, params)
        rows = cursor.fetchall()
        try:
            connection.rollback()
        except Exception:
            pass
        return rows
    finally:
        connection.close()


def build_community_tax_certificate_rows(date_from, date_to):
    sql = f"""
        SELECT
            p.PAYMENTDATE,
            COALESCE(NULLIF(TRIM(c.CTCNO), ''), TRIM(p.RECEIPTNO)) AS CTC_NO,
            COALESCE(NULLIF(TRIM(tp.OWNERNAME), ''), NULLIF(TRIM(p.PAIDBY), ''), '-') AS TAXPAYER_NAME,
            c.BASICTAXDUE,
            COALESCE(c.BUSTAXDUE, 0) + COALESCE(c.SALTAXDUE, 0) + COALESCE(c.RPTAXDUE, 0) AS AMOUNT,
            c.INTEREST,
            COALESCE(c.TOTALAMOUNTPAID, pd.AMOUNTPAID, 0) AS TOTAL_AMOUNT,
            COALESCE(NULLIF(TRIM(p.COLLECTOR), ''), NULLIF(TRIM(p.USERID), ''), 'UNSPECIFIED') AS CASHIER_NAME
        FROM PAYMENT p
        JOIN PAYMENTDETAIL pd ON pd.PAYMENT_ID = p.PAYMENT_ID
        LEFT JOIN COMMUNITYTAXCERTIFICATE c ON c.CTC_ID = pd.SOURCEID
        LEFT JOIN TAXPAYER tp ON tp.LOCAL_TIN = c.LOCAL_TIN
        WHERE p.PAYMENTDATE >= CAST(? AS DATE)
          AND p.PAYMENTDATE < DATEADD(1 DAY TO CAST(? AS DATE))
          {PAID_PAYMENT_SQL}
          AND (pd.SOURCE_CT IN ('CTCI', 'CTCC') OR pd.ITAXTYPE_CT = 'CTC')
        ORDER BY p.PAYMENTDATE, TRIM(p.RECEIPTNO), p.PAYMENT_ID
    """
    rows = [["DATE", "CTC NO.", "NAME", "BASIC", "AMOUNT", "INTEREST", "TOTAL", "CASHER"]]
    for row in database_rows(sql, (date_from, date_to)):
        rows.append(list(row))
    return rows


def write_community_tax_certificate_workbook(rows, output_path, date_from, date_to):
    workbook = load_workbook(TEMPLATE_DIR / "ABSTRACT_OF_COMMUNITY_TAX_CERTIFICATE_TEMPLATE.xlsx")
    sheet = workbook.active
    start_row = 4
    for index, row_values in enumerate(rows[1:], start=start_row):
        for col_index, value in enumerate(row_values[:8], start=1):
            cell = sheet.cell(index, col_index)
            cell.value = excel_value(value)
            if col_index in (4, 5, 6, 7):
                cell.number_format = '#,##0.00'
    output_path = save_workbook_with_fallback(workbook, output_path.with_suffix(".xlsx"))
    return output_path, len(rows) - 1 if rows else 0


def write_native_template_workbook(report_number, date_from, date_to, output_dir):
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / safe_filename(f"report_{report_number}_{date_from}_to_{date_to}.xlsx")

    if report_number == 25:
        rows = build_rpt_record_rows_from_fdb(date_from, date_to, None, None)
        row_count, output_path = write_rpt_record_workbook(rows, output_path, date_from, date_to)
        return output_path, row_count

    if report_number == 29:
        rows, daily_rows = build_abstract_general_collections_rows_from_fdb(date_from, date_to, None, None)
        row_count, output_path = write_abstract_general_collections_workbook(
            rows,
            daily_rows,
            output_path,
            date_from,
            date_to,
        )
        return output_path, row_count

    if report_number == 30:
        rows, daily_rows = build_abstract_trust_funds_rows_from_fdb(date_from, date_to, None, None)
        row_count, output_path = write_abstract_trust_funds_workbook(
            rows,
            daily_rows,
            output_path,
            date_from,
            date_to,
        )
        return output_path, row_count

    if report_number == 39:
        rows = build_community_tax_certificate_rows(date_from, date_to)
        output_path, row_count = write_community_tax_certificate_workbook(
            rows,
            output_path,
            date_from,
            date_to,
        )
        return output_path, row_count

    raise ValueError(f"Native template export for report {report_number} is not implemented.")


def delegated_parent_export(report_number, date_from, date_to, output_dir, collector=None):
    if not PARENT_COLLECTION_RUNNER.exists():
        searched = "; ".join(str(path) for path in parent_runner_candidates())
        raise FileNotFoundError(
            f"Parent collection runner was not found: {PARENT_COLLECTION_RUNNER}. Searched: {searched}"
        )

    command = [
        sys.executable,
        str(PARENT_COLLECTION_RUNNER),
        str(report_number),
        date_from,
        date_to,
        "--user",
        os.environ.get("FIREBIRD_USER", ""),
        "--password",
        os.environ.get("FIREBIRD_PASSWORD", ""),
    ]
    if collector:
        command.extend(["--collector", collector])
    result = subprocess.run(
        command,
        capture_output=True,
        text=True,
        timeout=180,
        check=False,
        cwd=str(PARENT_COLLECTION_RUNNER.parent),
    )
    output = "\n".join(part for part in (result.stdout, result.stderr) if part).strip()

    if result.returncode != 0:
        raise RuntimeError(output or f"Parent collection runner failed with exit code {result.returncode}.")

    match = re.search(r"Output file:\s*(.+)", output)
    if not match:
        raise RuntimeError(f"Parent collection runner did not return an output path. Output: {output}")

    source_path = Path(match.group(1).strip())
    if not source_path.exists():
        raise FileNotFoundError(f"Parent collection runner output was not found: {source_path}")

    rows_match = re.search(r"Rows exported:\s*(\d+)", output)
    row_count = int(rows_match.group(1)) if rows_match else 0

    output_dir.mkdir(parents=True, exist_ok=True)
    destination = output_dir / source_path.name
    if destination.exists():
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        destination = destination.with_name(f"{destination.stem}_{timestamp}{destination.suffix}")

    shutil.copy2(source_path, destination)
    return destination, row_count


def main():
    parser = argparse.ArgumentParser(description="Read-only Firebird Excel report exporter.")
    parser.add_argument("report_number", type=int)
    parser.add_argument("--date-from", required=True)
    parser.add_argument("--date-to", required=True)
    parser.add_argument("--output-dir", required=True)
    parser.add_argument("--collector", help="Optional collector filter for collector reports.")
    args = parser.parse_args()

    try:
        if args.report_number == OFFICIAL_BREAKDOWN_REPORT:
            output_path, row_count, totals = write_official_breakdown_workbook(
                args.date_from,
                args.date_to,
                Path(args.output_dir),
            )
            print(json.dumps({
                "ok": True,
                "mode": "read_only_excel_export",
                "report_number": args.report_number,
                "date_from": args.date_from,
                "date_to": args.date_to,
                "row_count": row_count,
                "rpt_gf": totals["rpt_gf"],
                "rpt_sf": totals["rpt_sf"],
                "rpt_municipal": totals["rpt_municipal"],
                "grand_total": totals["grand_total"],
                "path": str(output_path),
                "filename": output_path.name,
            }, default=scalar))
            return 0

        if args.report_number == ESRE_QUARTERLY_REPORT:
            output_path, row_count, grand_total = write_esre_quarterly_workbook(
                args.date_from,
                args.date_to,
                Path(args.output_dir),
            )
            print(json.dumps({
                "ok": True,
                "mode": "read_only_excel_export",
                "report_number": args.report_number,
                "date_from": args.date_from,
                "date_to": args.date_to,
                "row_count": row_count,
                "grand_total": grand_total,
                "path": str(output_path),
                "filename": output_path.name,
            }, default=scalar))
            return 0

        if args.report_number in RECEIPT_EXCEPTION_REPORTS:
            output_path, row_count = write_receipt_exception_workbook(
                args.report_number,
                args.date_from,
                args.date_to,
                Path(args.output_dir),
            )
            print(json.dumps({
                "ok": True,
                "mode": "read_only_excel_export",
                "report_number": args.report_number,
                "date_from": args.date_from,
                "date_to": args.date_to,
                "row_count": row_count,
                "path": str(output_path),
                "filename": output_path.name,
            }, default=scalar))
            return 0

        if args.report_number == COLLECTOR_RECEIPT_REPORT:
            output_path, row_count, grand_total = write_collector_receipt_workbook(
                args.date_from,
                args.date_to,
                Path(args.output_dir),
                args.collector,
            )
            print(json.dumps({
                "ok": True,
                "mode": "read_only_excel_export",
                "report_number": args.report_number,
                "date_from": args.date_from,
                "date_to": args.date_to,
                "row_count": row_count,
                "grand_total": grand_total,
                "path": str(output_path),
                "filename": output_path.name,
            }, default=scalar))
            return 0

        if args.report_number in NATIVE_TEMPLATE_REPORTS:
            output_path, row_count = write_native_template_workbook(
                args.report_number,
                args.date_from,
                args.date_to,
                Path(args.output_dir),
            )
            print(json.dumps({
                "ok": True,
                "mode": "read_only_excel_export",
                "report_number": args.report_number,
                "date_from": args.date_from,
                "date_to": args.date_to,
                "row_count": row_count,
                "path": str(output_path),
                "filename": output_path.name,
            }, default=scalar))
            return 0

        if args.report_number in PARENT_DELEGATED_REPORTS:
            output_path, row_count = delegated_parent_export(
                args.report_number,
                args.date_from,
                args.date_to,
                Path(args.output_dir),
                args.collector,
            )
            print(json.dumps({
                "ok": True,
                "mode": "read_only_excel_export",
                "report_number": args.report_number,
                "date_from": args.date_from,
                "date_to": args.date_to,
                "row_count": row_count,
                "path": str(output_path),
                "filename": output_path.name,
            }, default=scalar))
            return 0

        if args.report_number not in TEMPLATE_MAP:
            raise ValueError(f"Excel download for report {args.report_number} is not implemented yet.")

        payload = build_report(args.report_number, args.date_from, args.date_to)
        output_path, row_count = write_summary_workbook(
            args.report_number,
            payload,
            args.date_from,
            args.date_to,
            Path(args.output_dir),
        )

        print(json.dumps({
            "ok": True,
            "mode": "read_only_excel_export",
            "report_number": args.report_number,
            "date_from": args.date_from,
            "date_to": args.date_to,
            "row_count": row_count,
            "path": str(output_path),
            "filename": output_path.name,
        }, default=scalar))
        return 0
    except Exception as exc:
        print(json.dumps({
            "ok": False,
            "mode": "read_only_excel_export",
            "error": str(exc),
            "error_type": exc.__class__.__name__,
        }))
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
