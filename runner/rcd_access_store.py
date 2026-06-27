import argparse
from copy import copy
import json
import os
import re
import shutil
import sys
import tempfile
from datetime import datetime
from pathlib import Path

import pyodbc
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DB_PATH = Path(os.environ.get("RCD_ACCESS_DB") or PROJECT_ROOT / "backend" / "database" / "rcd" / "rcd_remittance.accdb").resolve()
TEMPLATE_PATH = PROJECT_ROOT / "template" / "RCD_UPDATED.xlsx"
ACCESS_DRIVER = "Microsoft Access Driver (*.mdb, *.accdb)"
COLLECTOR_FULL_NAMES = {
    "FLORA MY": "FLORA MY D. FERRER",
    "AGNES": "AGNES B. ELLO",
    "RICARDO": "RICARDO T. ENOPIA",
    "IRIS": "ANGELIQUE IRIS A. RAFALES",
    "EMILY": "EMILY E. CREDO",
}


def emit(payload, code=0):
    print(json.dumps(payload, ensure_ascii=False, default=str))
    raise SystemExit(code)


def digits(value):
    return re.sub(r"\D", "", str(value or ""))


def count_range(start, end):
    start_num = int(digits(start) or 0)
    end_num = int(digits(end) or 0)
    if not start_num or not end_num or end_num < start_num:
        return 0
    return end_num - start_num + 1


def money(value):
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def collector_full_name(value):
    key = str(value or "").strip().upper()
    return COLLECTOR_FULL_NAMES.get(key, str(value or "").strip().upper())


def safe_filename(value):
    return re.sub(r"_+", "_", re.sub(r"[^A-Za-z0-9-]+", "_", str(value or "").strip())).strip("_")


def form_type_label(value):
    return "Comm Tax." if str(value or "").strip() == "Community Tax Certificate" else str(value or "")


def connect():
    if not DB_PATH.exists():
        emit({"ok": False, "error": f"AccessDB file was not found: {DB_PATH}"}, 1)

    conn = pyodbc.connect(f"DRIVER={{{ACCESS_DRIVER}}};DBQ={DB_PATH};")
    ensure_schema(conn)
    return conn


def table_names(cursor):
    return {row.table_name.lower() for row in cursor.tables(tableType="TABLE")}


def column_names(cursor, table):
    return {row.column_name.lower() for row in cursor.columns(table=table)}


def try_execute(cursor, sql):
    try:
        cursor.execute(sql)
    except pyodbc.Error:
        pass


def ensure_column(cursor, existing_columns, table, column, definition):
    if column in existing_columns:
        return
    try_execute(cursor, f"ALTER TABLE {table} ADD COLUMN {column} {definition}")


def ensure_schema(conn):
    cursor = conn.cursor()
    tables = table_names(cursor)

    if "rcd_accountability_snapshots" not in tables:
        cursor.execute(
            """
            CREATE TABLE rcd_accountability_snapshots (
                id COUNTER PRIMARY KEY,
                batch_id INTEGER NOT NULL,
                form_type TEXT(80) NOT NULL,
                beginning_qty INTEGER,
                beginning_from TEXT(30),
                beginning_to TEXT(30),
                receipt_qty INTEGER,
                receipt_from TEXT(30),
                receipt_to TEXT(30),
                issued_qty INTEGER,
                issued_from TEXT(30),
                issued_to TEXT(30),
                ending_qty INTEGER,
                ending_from TEXT(30),
                ending_to TEXT(30),
                created_at DATETIME
            )
            """
        )

    batch_cols = column_names(cursor, "rcd_batches")
    if "report_no" not in batch_cols:
        try_execute(cursor, "ALTER TABLE rcd_batches ADD COLUMN report_no TEXT(80)")
    remittance_columns = {
        "remittance_status": "TEXT(40)",
        "remitted_by": "TEXT(100)",
        "remitted_at": "DATETIME",
        "remitted_to_aco_by": "TEXT(100)",
        "remitted_to_aco_at": "DATETIME",
        "received_by": "TEXT(100)",
        "received_at": "DATETIME",
        "received_by_aco": "TEXT(100)",
        "received_by_aco_at": "DATETIME",
        "amount_remitted": "CURRENCY",
        "amount_received": "CURRENCY",
        "cash_amount": "CURRENCY",
        "check_amount": "CURRENCY",
        "variance_amount": "CURRENCY",
        "reference_no": "TEXT(100)",
        "remittance_remarks": "MEMO",
        "created_by": "TEXT(100)",
        "updated_by": "TEXT(100)",
        "printed_by": "TEXT(100)",
        "printed_at": "DATETIME",
        "voided_by": "TEXT(100)",
        "voided_at": "DATETIME",
        "void_reason": "MEMO",
    }
    for column, definition in remittance_columns.items():
        ensure_column(cursor, batch_cols, "rcd_batches", column, definition)

    line_cols = column_names(cursor, "rcd_collection_lines")
    ensure_column(cursor, line_cols, "rcd_collection_lines", "raw_json", "MEMO")

    conn.commit()


def normalize_payload(payload):
    form = payload.get("form") or {}
    lines = payload.get("lines") or []
    report_no = (payload.get("report_no") or form.get("reportNo") or form.get("report_no") or "").strip()

    return {
        "lookup_key": payload.get("lookup_key") or "",
        "report_no": report_no,
        "report_date": form.get("collectionDate") or form.get("collection_date") or datetime.now().date().isoformat(),
        "collector": form.get("collector") or "",
        "template": "100_GF + 200_SEF",
        "status": payload.get("status") or "Draft",
        "lines": lines,
    }


def find_batch_id(cursor, key):
    key = str(key or "").strip()
    if key.startswith("__dbid:"):
        row = cursor.execute("SELECT id FROM rcd_batches WHERE id = ?", int(key.replace("__dbid:", "", 1))).fetchone()
        return int(row.id) if row else None
    if not key:
        return None
    row = cursor.execute("SELECT id FROM rcd_batches WHERE report_no = ?", key).fetchone()
    return int(row.id) if row else None


def find_blank_report_batch_id(cursor, data):
    if data["lookup_key"] or data["report_no"]:
        return None

    row = cursor.execute(
        """
        SELECT TOP 1 id
          FROM rcd_batches
         WHERE (report_no IS NULL OR report_no = '')
           AND report_date = ?
           AND collector = ?
         ORDER BY id DESC
        """,
        data["report_date"],
        data["collector"],
    ).fetchone()

    return int(row.id) if row else None


def log_transaction(cursor, batch_id, action, details):
    cursor.execute(
        """
        INSERT INTO rcd_access_audit_logs
            (batch_id, log_action, performed_by, details, created_at)
        VALUES (?, ?, ?, ?, ?)
        """,
        batch_id,
        action,
        "LGU Treasury System",
        json.dumps(details, ensure_ascii=False),
        datetime.now(),
    )


def status_key(value):
    return str(value or "").strip().lower()


def deletable_status(status):
    return status_key(status) == "draft"


def remittable_status(status):
    return status_key(status) in ("saved", "for remittance", "ready for remittance")


def locked_status(status):
    return status_key(status) in ("saved", "printed", "remitted", "remitted to aco", "received by aco")


def line_or_numbers(line):
    start = int(digits(line.get("receiptFrom") or line.get("receipt_from") or "") or 0)
    end = int(digits(line.get("receiptTo") or line.get("receipt_to") or line.get("receiptFrom") or line.get("receipt_from") or "") or 0)
    if not start or not end or end < start or end - start > 5000:
        return []
    return list(range(start, end + 1))


def validate_batch_for_remittance(batch, cursor=None):
    errors = []
    warnings = []
    lines = batch.get("lines") or []
    if not lines:
        errors.append("RCD has no OR records.")
    if money(batch.get("total")) <= 0:
        errors.append("Total collection amount must be greater than zero.")

    seen = set()
    duplicate_numbers = set()
    for line in lines:
        validation_status = status_key(line.get("validationStatus") or line.get("validation_status"))
        if validation_status in ("void", "cancelled", "canceled"):
            errors.append(f"OR range {line.get('receiptFrom')} to {line.get('receiptTo')} contains {validation_status} receipt(s).")
        for number in line_or_numbers(line):
            if number in seen:
                duplicate_numbers.add(number)
            seen.add(number)

    if duplicate_numbers:
        preview = ", ".join(str(number) for number in sorted(duplicate_numbers)[:10])
        errors.append(f"Duplicate OR number(s) found: {preview}.")

    # Accountable form assignment is only enforceable once releases are encoded.
    # For now we retain this as a warning instead of blocking old/manual batches.
    if not batch.get("collector"):
        warnings.append("Collector is blank; accountable form assignment cannot be verified.")
    elif cursor is not None:
        collector = str(batch.get("collector") or "").strip().upper()
        for line in lines:
            form_type = form_type_label(line.get("formType") or line.get("form_type") or "")
            receipt_from = digits(line.get("receiptFrom") or line.get("receipt_from") or "")
            receipt_to = digits(line.get("receiptTo") or line.get("receipt_to") or receipt_from)
            if not receipt_from or not receipt_to:
                continue
            try:
                release = cursor.execute(
                    """
                    SELECT TOP 1 collector
                      FROM rcd_accountable_form_releases
                     WHERE form_type = ?
                       AND VAL(receipt_no_from) <= ?
                       AND VAL(receipt_no_to) >= ?
                     ORDER BY released_at DESC, id DESC
                    """,
                    form_type,
                    int(receipt_from),
                    int(receipt_to),
                ).fetchone()
            except pyodbc.Error:
                warnings.append("Accountable form assignment table is not ready; collector assignment was not enforced.")
                continue
            if not release:
                warnings.append(f"No accountable form release record found for {form_type} OR {receipt_from}-{receipt_to}.")
            elif str(release.collector or "").strip().upper() != collector:
                errors.append(f"{form_type} OR {receipt_from}-{receipt_to} is assigned to {release.collector}, not {batch.get('collector')}.")

    return errors, warnings


def save_batch(payload):
    data = normalize_payload(payload)
    lines = data["lines"]
    if not lines:
        emit({"ok": False, "error": "No RCD collection lines were provided."}, 1)

    saved_total = sum(money(line.get("collectorAmount") or line.get("collector_amount")) for line in lines)
    fdb_total = sum(money(line.get("fdbAmount") or line.get("fdb_amount")) for line in lines)
    difference = saved_total - fdb_total
    receipt_from = next((line.get("receiptFrom") or line.get("receipt_from") for line in lines if line.get("receiptFrom") or line.get("receipt_from")), "")
    receipt_to = next((line.get("receiptTo") or line.get("receipt_to") or receipt_from for line in reversed(lines) if line.get("receiptTo") or line.get("receipt_to") or line.get("receiptFrom") or line.get("receipt_from")), "")
    now = datetime.now()

    conn = connect()
    cursor = conn.cursor()
    batch_id = find_batch_id(cursor, data["lookup_key"] or data["report_no"]) or find_blank_report_batch_id(cursor, data)

    was_update = bool(batch_id)
    if batch_id:
        cursor.execute(
            """
            UPDATE rcd_batches
               SET report_date = ?, collector = ?, template_code = ?, fund_label = ?,
                   receipt_no_from = ?, receipt_no_to = ?, fdb_total = ?, saved_total = ?,
                   difference = ?, status = ?, updated_at = ?
             WHERE id = ?
            """,
            data["report_date"],
            data["collector"],
            data["template"],
            data["template"],
            receipt_from,
            receipt_to,
            fdb_total,
            saved_total,
            difference,
            data["status"],
            now,
            batch_id,
        )
        cursor.execute("DELETE FROM rcd_collection_lines WHERE batch_id = ?", batch_id)
        cursor.execute("DELETE FROM rcd_accountability_snapshots WHERE batch_id = ?", batch_id)
    else:
        cursor.execute(
            """
            INSERT INTO rcd_batches
                (report_no, report_date, collector, template_code, fund_label,
                 receipt_no_from, receipt_no_to, fdb_total, saved_total, difference,
                 status, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            data["report_no"],
            data["report_date"],
            data["collector"],
            data["template"],
            data["template"],
            receipt_from,
            receipt_to,
            fdb_total,
            saved_total,
            difference,
            data["status"],
            now,
            now,
        )
        batch_id = int(cursor.execute("SELECT @@IDENTITY AS id").fetchone().id)

    for index, line in enumerate(lines, start=1):
        form_type = form_type_label(line.get("formType") or line.get("form_type") or "")
        line_from = line.get("receiptFrom") or line.get("receipt_from") or ""
        line_to = line.get("receiptTo") or line.get("receipt_to") or line_from
        collector_amount = money(line.get("collectorAmount") or line.get("collector_amount"))
        line_fdb_total = money(line.get("fdbAmount") or line.get("fdb_amount"))
        line_difference = collector_amount - line_fdb_total
        receipt_count = count_range(line_from, line_to)

        cursor.execute(
            """
            INSERT INTO rcd_collection_lines
                (batch_id, line_no, form_type, receipt_no_from, receipt_no_to, receipt_count,
                 fdb_total, saved_total, difference, validation_status, raw_json, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            batch_id,
            index,
            form_type,
            line_from,
            line_to,
            receipt_count,
            line_fdb_total,
            collector_amount,
            line_difference,
            line.get("validationStatus") or line.get("validation_status") or "Not validated",
            json.dumps(line, ensure_ascii=False),
            now,
        )

        cursor.execute(
            """
            INSERT INTO rcd_accountability_snapshots
                (batch_id, form_type, beginning_qty, beginning_from, beginning_to,
                 receipt_qty, receipt_from, receipt_to, issued_qty, issued_from, issued_to,
                 ending_qty, ending_from, ending_to, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            batch_id,
            form_type,
            int(line.get("beginningQty") or 0),
            line.get("beginningFrom") or "",
            line.get("beginningTo") or "",
            int(line.get("receiptAccountQty") or 0),
            line.get("receiptAccountFrom") or "",
            line.get("receiptAccountTo") or "",
            receipt_count,
            line_from,
            line_to,
            int(line.get("endingQty") or 0),
            line.get("endingFrom") or "",
            line.get("endingTo") or "",
            now,
        )

    log_transaction(cursor, batch_id, "RCD_UPDATED" if was_update else "RCD_CREATED", {
        "report_no": data["report_no"],
        "collector": data["collector"],
        "report_date": data["report_date"],
        "status": data["status"],
        "line_count": len(lines),
        "saved_total": saved_total,
        "fdb_total": fdb_total,
        "difference": difference,
    })

    conn.commit()
    emit({"ok": True, "data": get_batch(f"__dbid:{batch_id}", conn=conn), "message": "RCD saved to AccessDB."})


def row_dict(row):
    return {cursor_description[0]: value for cursor_description, value in zip(row.cursor_description, row)}


def list_batches():
    conn = connect()
    cursor = conn.cursor()
    rows = cursor.execute(
        """
        SELECT b.id, b.report_no, b.report_date, b.collector, b.template_code, b.saved_total,
               b.status, b.created_at, b.updated_at, b.receipt_no_from, b.receipt_no_to,
               b.amount_remitted, b.amount_received, b.variance_amount,
               b.remitted_to_aco_at, b.received_by_aco_at,
               COUNT(l.id) AS entries,
               SUM(l.receipt_count) AS receipt_count
          FROM rcd_batches AS b
          LEFT JOIN rcd_collection_lines AS l ON b.id = l.batch_id
         GROUP BY b.id, b.report_no, b.report_date, b.collector, b.template_code, b.saved_total,
                  b.status, b.created_at, b.updated_at, b.receipt_no_from, b.receipt_no_to,
                  b.amount_remitted, b.amount_received, b.variance_amount,
                  b.remitted_to_aco_at, b.received_by_aco_at
         ORDER BY b.report_date DESC, b.id DESC
        """
    ).fetchall()

    data = []
    for row in rows:
        forms = cursor.execute(
            "SELECT DISTINCT form_type FROM rcd_collection_lines WHERE batch_id = ?",
            row.id,
        ).fetchall()
        data.append({
            "db_id": row.id,
            "action_key": row.report_no or f"__dbid:{row.id}",
            "id": row.report_no or "-",
            "date": str(row.report_date)[:10],
            "collector": row.collector,
            "fund": row.template_code,
            "forms": " / ".join(form.form_type for form in forms),
            "entries": int(row.entries or 0),
            "receipt_count": int(row.receipt_count or 0),
            "receipt_no_from": row.receipt_no_from or "",
            "receipt_no_to": row.receipt_no_to or "",
            "total": money(row.saved_total),
            "amount_remitted": money(row.amount_remitted),
            "amount_received": money(row.amount_received),
            "variance_amount": money(row.variance_amount),
            "stage": row.status,
            "can_remit": remittable_status(row.status),
            "can_delete": deletable_status(row.status),
            "created_at": str(row.created_at) if row.created_at else "",
            "updated_at": str(row.updated_at) if row.updated_at else "",
            "remitted_to_aco_at": str(row.remitted_to_aco_at) if row.remitted_to_aco_at else "",
            "received_by_aco_at": str(row.received_by_aco_at) if row.received_by_aco_at else "",
        })

    emit({"ok": True, "data": data})


def get_batch(report_no, conn=None, emit_result=True):
    own_conn = conn is None
    conn = conn or connect()
    cursor = conn.cursor()
    batch_id = find_batch_id(cursor, report_no)
    batch = cursor.execute("SELECT * FROM rcd_batches WHERE id = ?", batch_id).fetchone() if batch_id else None
    if not batch:
        if own_conn and emit_result:
            emit({"ok": False, "error": "RCD report was not found."}, 1)
        return None

    cols = [col[0] for col in cursor.description]
    batch_data = dict(zip(cols, batch))
    lines = cursor.execute(
        "SELECT * FROM rcd_collection_lines WHERE batch_id = ? ORDER BY line_no",
        batch_data["id"],
    ).fetchall()

    line_cols = [col[0] for col in cursor.description]
    parsed_lines = []
    for line in lines:
        record = dict(zip(line_cols, line))
        raw = {}
        if record.get("raw_json"):
            try:
                raw = json.loads(record["raw_json"])
            except json.JSONDecodeError:
                raw = {}
        raw.update({
            "formType": form_type_label(raw.get("formType") or record.get("form_type")),
            "receiptFrom": raw.get("receiptFrom") or record.get("receipt_no_from"),
            "receiptTo": raw.get("receiptTo") or record.get("receipt_no_to"),
            "collectorAmount": money(raw.get("collectorAmount") or record.get("saved_total")),
            "fdbAmount": money(raw.get("fdbAmount") or record.get("fdb_total")),
            "validationStatus": raw.get("validationStatus") or record.get("validation_status"),
            "validated": bool(raw.get("validated", True)),
        })
        parsed_lines.append(raw)

    result = {
        "db_id": batch_data["id"],
        "action_key": batch_data.get("report_no") or f"__dbid:{batch_data['id']}",
        "id": batch_data.get("report_no") or "-",
        "date": str(batch_data.get("report_date"))[:10],
        "collector": batch_data.get("collector"),
        "fund": batch_data.get("template_code"),
        "status": batch_data.get("status"),
        "total": money(batch_data.get("saved_total")),
        "fdb_total": money(batch_data.get("fdb_total")),
        "difference": money(batch_data.get("difference")),
        "remittance_status": batch_data.get("remittance_status") or "",
        "remitted_by": batch_data.get("remitted_by") or "",
        "remitted_at": str(batch_data.get("remitted_at")) if batch_data.get("remitted_at") else "",
        "remitted_to_aco_by": batch_data.get("remitted_to_aco_by") or "",
        "remitted_to_aco_at": str(batch_data.get("remitted_to_aco_at")) if batch_data.get("remitted_to_aco_at") else "",
        "received_by": batch_data.get("received_by") or "",
        "received_at": str(batch_data.get("received_at")) if batch_data.get("received_at") else "",
        "received_by_aco": batch_data.get("received_by_aco") or "",
        "received_by_aco_at": str(batch_data.get("received_by_aco_at")) if batch_data.get("received_by_aco_at") else "",
        "amount_remitted": money(batch_data.get("amount_remitted")),
        "amount_received": money(batch_data.get("amount_received")),
        "cash_amount": money(batch_data.get("cash_amount")),
        "check_amount": money(batch_data.get("check_amount")),
        "variance_amount": money(batch_data.get("variance_amount")),
        "reference_no": batch_data.get("reference_no") or "",
        "remittance_remarks": batch_data.get("remittance_remarks") or "",
        "form": {
            "reportNo": batch_data.get("report_no") or "",
            "collectionDate": str(batch_data.get("report_date"))[:10],
            "collector": batch_data.get("collector") or "",
            "template": batch_data.get("template_code") or "100_GF",
        },
        "lines": parsed_lines,
    }

    if own_conn and emit_result:
        emit({"ok": True, "data": result})
    return result


def copy_row_style(sheet, source_row, target_row, max_col=13):
    sheet.row_dimensions[target_row].height = sheet.row_dimensions[source_row].height
    for col in range(1, max_col + 1):
        source = sheet.cell(source_row, col)
        target = sheet.cell(target_row, col)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.alignment:
            target.alignment = copy(source.alignment)
        if source.font:
            target.font = copy(source.font)
        if source.fill:
            target.fill = copy(source.fill)
        if source.border:
            target.border = copy(source.border)


def merge_collection_row(sheet, row):
    for range_text in (f"A{row}:C{row}", f"D{row}:F{row}", f"G{row}:I{row}", f"J{row}:M{row}"):
        try:
            sheet.merge_cells(range_text)
        except ValueError:
            pass


def shift_range(range_text, row_index, amount):
    min_col, min_row, max_col, max_row = range_boundaries(range_text)
    if min_row >= row_index:
        min_row += amount
        max_row += amount
    elif max_row >= row_index:
        max_row += amount
    return f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"


def insert_rows_preserve_merges(sheet, row_index, amount):
    merged_ranges = [str(merged_range) for merged_range in sheet.merged_cells.ranges]
    for merged_range in merged_ranges:
        sheet.unmerge_cells(merged_range)
    sheet.insert_rows(row_index, amount)
    for merged_range in merged_ranges:
        sheet.merge_cells(shift_range(merged_range, row_index, amount))


def unmerge_rows(sheet, start_row, end_row):
    for merged_range in list(sheet.merged_cells.ranges):
        if merged_range.max_row >= start_row and merged_range.min_row <= end_row:
            sheet.unmerge_cells(str(merged_range))


def clear_row_values(sheet, row, max_col=13):
    for col in range(1, max_col + 1):
        cell = sheet.cell(row, col)
        if cell.__class__.__name__ == "MergedCell":
            continue
        cell.value = None


def template_date(value):
    try:
        return datetime.fromisoformat(str(value)[:10])
    except ValueError:
        return str(value or "")


def is_sef_line(line):
    form_type = form_type_label(line.get("formType") or line.get("form_type")).upper()
    return "AF 56" in form_type or "RPT" in form_type or "SEF" in form_type


def fill_rcd_sheet(sheet, batch, lines, fund_code):
    sheet["A4"] = "Fund: 100 General Fund" if fund_code == "100_GF" else "Fund: 200 Special Education Fund"
    sheet["J4"] = template_date(batch["date"])
    sheet["J4"].number_format = "mmmm dd, yyyy"
    officer_name = collector_full_name(batch["collector"])
    sheet["C5"] = officer_name
    sheet["J5"] = "" if batch["id"] == "-" else batch["id"]

    collection_start = 12
    base_collection_rows = 2
    extra_collection_rows = max(len(lines) - base_collection_rows, 0)
    if extra_collection_rows:
        insert_rows_preserve_merges(sheet, 14, extra_collection_rows)
        for row in range(14, 14 + extra_collection_rows):
            copy_row_style(sheet, 13, row)
            merge_collection_row(sheet, row)

    collection_total_row = 14 + extra_collection_rows
    unmerge_rows(sheet, collection_start, collection_total_row - 1)
    for row in range(collection_start, collection_total_row):
        clear_row_values(sheet, row)
        merge_collection_row(sheet, row)

    for offset, line in enumerate(lines):
        row = collection_start + offset
        sheet[f"A{row}"] = form_type_label(line.get("formType"))
        sheet[f"D{row}"] = line.get("receiptFrom") or ""
        sheet[f"G{row}"] = line.get("receiptTo") or line.get("receiptFrom") or ""
        sheet[f"J{row}"] = money(line.get("collectorAmount"))
        sheet[f"J{row}"].number_format = "#,##0.00"

    sheet[f"J{collection_total_row}"] = f"=SUM(J{collection_start}:J{collection_total_row - 1})"

    accountability_start = 26 + extra_collection_rows
    accountability_summary_row = 28 + extra_collection_rows
    base_accountability_rows = 2
    extra_accountability_rows = max(len(lines) - base_accountability_rows, 0)
    if extra_accountability_rows:
        insert_rows_preserve_merges(sheet, accountability_summary_row, extra_accountability_rows)
        for row in range(accountability_summary_row, accountability_summary_row + extra_accountability_rows):
            copy_row_style(sheet, accountability_summary_row - 1, row)

    accountability_end = accountability_summary_row + extra_accountability_rows
    unmerge_rows(sheet, accountability_start, accountability_end - 1)
    for row in range(accountability_start, accountability_end):
        clear_row_values(sheet, row)

    for offset, line in enumerate(lines):
        row = accountability_start + offset
        issued_qty = count_range(line.get("receiptFrom"), line.get("receiptTo"))
        sheet[f"A{row}"] = form_type_label(line.get("formType"))
        sheet[f"B{row}"] = int(line.get("beginningQty") or 0)
        sheet[f"C{row}"] = line.get("beginningFrom") or ""
        sheet[f"D{row}"] = line.get("beginningTo") or ""
        sheet[f"E{row}"] = int(line.get("receiptAccountQty") or 0)
        sheet[f"F{row}"] = line.get("receiptAccountFrom") or ""
        sheet[f"G{row}"] = line.get("receiptAccountTo") or ""
        sheet[f"H{row}"] = issued_qty
        sheet[f"I{row}"] = line.get("receiptFrom") or ""
        sheet[f"J{row}"] = line.get("receiptTo") or line.get("receiptFrom") or ""
        sheet[f"K{row}"] = int(line.get("endingQty") or 0)
        sheet[f"L{row}"] = line.get("endingFrom") or ""
        sheet[f"M{row}"] = line.get("endingTo") or ""

    summary_collection_row = 30 + extra_collection_rows + extra_accountability_rows
    sheet[f"J{summary_collection_row}"] = f"=J{collection_total_row}"
    signature_row = 38 + extra_collection_rows + extra_accountability_rows
    sheet[f"A{signature_row}"] = officer_name


def export_batch(report_no):
    batch = get_batch(report_no, emit_result=False)
    if not batch:
        emit({"ok": False, "error": "RCD report was not found."}, 1)

    if not TEMPLATE_PATH.exists():
        emit({"ok": False, "error": f"RCD template was not found: {TEMPLATE_PATH}"}, 1)

    out_dir = Path(tempfile.gettempdir()) / "lgu_treasury_rcd"
    out_dir.mkdir(parents=True, exist_ok=True)
    filename_key = f"{collector_full_name(batch['collector'])}_{batch['date']}"
    filename = f"{safe_filename(filename_key)}.xlsx"
    path = out_dir / filename
    shutil.copy2(TEMPLATE_PATH, path)

    workbook = load_workbook(path)
    if "E_GOV_ONLINE_PAYMENT" in workbook.sheetnames:
        workbook.remove(workbook["E_GOV_ONLINE_PAYMENT"])

    lines = batch["lines"] or []
    fill_rcd_sheet(workbook["100_GF"], batch, lines, "100_GF")
    fill_rcd_sheet(workbook["200_SEF"], batch, [line for line in lines if is_sef_line(line)], "200_SEF")

    workbook.save(path)

    conn = connect()
    cursor = conn.cursor()
    batch_id = find_batch_id(cursor, report_no)
    if batch_id:
        cursor.execute("UPDATE rcd_batches SET status = ?, updated_at = ? WHERE id = ?", "Saved", datetime.now(), batch_id)
        log_transaction(cursor, batch_id, "RCD_EXPORTED", {
            "report_no": report_no,
            "filename": filename,
            "sheets": ["100_GF", "200_SEF"],
            "status": "Saved",
        })
        conn.commit()

    emit({"ok": True, "path": str(path), "filename": filename})


def delete_batch(report_no):
    conn = connect()
    cursor = conn.cursor()
    batch_id = find_batch_id(cursor, report_no)
    if not batch_id:
        emit({"ok": False, "error": "RCD report was not found."}, 1)

    batch = cursor.execute("SELECT report_no, collector, report_date, saved_total, status FROM rcd_batches WHERE id = ?", batch_id).fetchone()
    if batch and not deletable_status(batch.status):
        emit({
            "ok": False,
            "error": f"Cannot delete an RCD with status '{batch.status}'. Use Void/Cancel with reason instead.",
        }, 1)

    log_transaction(cursor, batch_id, "RCD_DELETED", {
        "report_no": batch.report_no if batch else "",
        "collector": batch.collector if batch else "",
        "report_date": str(batch.report_date) if batch else "",
        "saved_total": money(batch.saved_total) if batch else 0,
        "status": batch.status if batch else "",
    })
    cursor.execute("DELETE FROM rcd_collection_lines WHERE batch_id = ?", batch_id)
    cursor.execute("DELETE FROM rcd_accountability_snapshots WHERE batch_id = ?", batch_id)
    cursor.execute("DELETE FROM rcd_remittance_events WHERE batch_id = ?", batch_id)
    cursor.execute("DELETE FROM rcd_batches WHERE id = ?", batch_id)
    conn.commit()

    emit({"ok": True, "message": "RCD batch deleted.", "deleted_id": batch_id})


def remit_batch(payload):
    lookup_key = payload.get("report_no") or payload.get("lookup_key") or ""
    batch = get_batch(lookup_key, emit_result=False)
    if not batch:
        emit({"ok": False, "error": "RCD report was not found."}, 1)

    if not remittable_status(batch.get("status")):
        emit({
            "ok": False,
            "error": f"RCD status must be Saved, For Remittance, or Ready for Remittance before remit. Current status: {batch.get('status') or '-'}",
        }, 1)

    conn = connect()
    cursor = conn.cursor()
    errors, warnings = validate_batch_for_remittance(batch, cursor)
    total = money(batch.get("total"))
    amount_remitted = money(payload.get("amount_remitted"))
    cash_amount = money(payload.get("cash_amount"))
    check_amount = money(payload.get("check_amount"))
    variance = round(amount_remitted - total, 2)
    reference_no = str(payload.get("reference_no") or "").strip()
    received_by = str(payload.get("received_by") or "").strip()
    remitted_by = str(payload.get("remitted_by") or "LGU Treasury System").strip()
    remarks = str(payload.get("remittance_remarks") or "").strip()
    received_at_value = payload.get("received_at") or datetime.now().isoformat(sep=" ", timespec="seconds")
    try:
        received_at = datetime.fromisoformat(str(received_at_value).replace("Z", "+00:00")).replace(tzinfo=None)
    except ValueError:
        received_at = datetime.now()

    if amount_remitted <= 0:
        errors.append("Amount remitted must be greater than zero.")
    if round(cash_amount + check_amount, 2) != round(amount_remitted, 2):
        errors.append("Cash amount plus check amount must equal amount remitted.")
    if variance != 0 and not remarks:
        errors.append("Variance requires remittance remarks.")
    if errors:
        emit({"ok": False, "error": "Remittance validation failed.", "errors": errors, "warnings": warnings}, 1)

    batch_id = find_batch_id(cursor, lookup_key)
    now = datetime.now()
    cursor.execute(
        """
        UPDATE rcd_batches
           SET status = ?, remittance_status = ?, remitted_by = ?, remitted_at = ?,
               remitted_to_aco_by = ?, remitted_to_aco_at = ?,
               received_by = ?, received_at = ?, amount_remitted = ?, cash_amount = ?,
               check_amount = ?, variance_amount = ?, reference_no = ?,
               remittance_remarks = ?, updated_at = ?
         WHERE id = ?
        """,
        "Remitted to ACO",
        "Remitted to ACO",
        remitted_by,
        now,
        remitted_by,
        now,
        received_by,
        received_at,
        amount_remitted,
        cash_amount,
        check_amount,
        variance,
        reference_no,
        remarks,
        now,
        batch_id,
    )
    cursor.execute(
        """
        INSERT INTO rcd_remittance_events
            (batch_id, event_type, event_at, performed_by, received_by, reference_no, amount, remarks, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        batch_id,
        "REMIT",
        now,
        remitted_by,
        received_by,
        reference_no,
        amount_remitted,
        remarks,
        now,
    )
    log_transaction(cursor, batch_id, "RCD_REMITTED", {
        "old_status": batch.get("status"),
        "new_status": "Remitted to ACO",
        "amount_remitted": amount_remitted,
        "cash_amount": cash_amount,
        "check_amount": check_amount,
        "variance_amount": variance,
        "reference_no": reference_no,
        "received_by": received_by,
        "warnings": warnings,
    })
    conn.commit()

    emit({"ok": True, "data": get_batch(f"__dbid:{batch_id}", conn=conn), "warnings": warnings, "message": "RCD successfully remitted."})


def receive_remittance(payload):
    lookup_key = payload.get("report_no") or payload.get("lookup_key") or ""
    batch = get_batch(lookup_key, emit_result=False)
    if not batch:
        emit({"ok": False, "error": "RCD report was not found."}, 1)
    if status_key(batch.get("status")) != "remitted to aco":
        emit({
            "ok": False,
            "error": f"RCD must be Remitted to ACO before receiving. Current status: {batch.get('status') or '-'}",
        }, 1)

    amount_received = money(payload.get("amount_received"))
    total = money(batch.get("total"))
    variance = round(total - amount_received, 2)
    new_status = "With Variance" if variance != 0 else "Received by ACO"
    received_by = str(payload.get("received_by_aco") or payload.get("received_by") or "").strip()
    remarks = str(payload.get("remittance_remarks") or "").strip()
    confirmed = bool(payload.get("confirmed"))
    received_at_value = payload.get("received_by_aco_at") or payload.get("received_at") or datetime.now().isoformat(sep=" ", timespec="seconds")
    try:
        received_at = datetime.fromisoformat(str(received_at_value).replace("Z", "+00:00")).replace(tzinfo=None)
    except ValueError:
        received_at = datetime.now()

    errors = []
    if amount_received <= 0:
        errors.append("Amount received must be greater than zero.")
    if not received_by:
        errors.append("Received by ACO is required.")
    if not confirmed:
        errors.append("Confirmation checkbox is required.")
    if variance != 0 and not remarks:
        errors.append("Variance requires remarks.")
    if errors:
        emit({"ok": False, "error": "Receive remittance validation failed.", "errors": errors}, 1)

    conn = connect()
    cursor = conn.cursor()
    batch_id = find_batch_id(cursor, lookup_key)
    now = datetime.now()
    cursor.execute(
        """
        UPDATE rcd_batches
           SET status = ?, remittance_status = ?, received_by = ?, received_at = ?,
               received_by_aco = ?, received_by_aco_at = ?, amount_received = ?,
               variance_amount = ?, remittance_remarks = ?, updated_at = ?
         WHERE id = ?
        """,
        new_status,
        new_status,
        received_by,
        received_at,
        received_by,
        received_at,
        amount_received,
        variance,
        remarks,
        now,
        batch_id,
    )
    cursor.execute(
        """
        INSERT INTO rcd_remittance_events
            (batch_id, event_type, event_at, performed_by, received_by, reference_no, amount, remarks, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        batch_id,
        "RECEIVE_REMITTANCE",
        now,
        received_by,
        received_by,
        batch.get("reference_no") or "",
        amount_received,
        remarks,
        now,
    )
    log_transaction(cursor, batch_id, "RCD_RECEIVED_BY_ACO", {
        "old_status": batch.get("status"),
        "new_status": new_status,
        "amount_received": amount_received,
        "variance_amount": variance,
        "received_by_aco": received_by,
    })
    conn.commit()
    emit({"ok": True, "data": get_batch(f"__dbid:{batch_id}", conn=conn), "message": "RCD remittance received by ACO."})


def audit_trail(report_no):
    conn = connect()
    cursor = conn.cursor()
    batch_id = find_batch_id(cursor, report_no)
    if not batch_id:
        emit({"ok": False, "error": "RCD report was not found."}, 1)
    rows = cursor.execute(
        """
        SELECT log_action, performed_by, details, created_at
          FROM rcd_access_audit_logs
         WHERE batch_id = ?
         ORDER BY created_at DESC, id DESC
        """,
        batch_id,
    ).fetchall()
    data = []
    for row in rows:
        try:
            details = json.loads(row.details) if row.details else {}
        except json.JSONDecodeError:
            details = row.details or ""
        data.append({
            "action": row.log_action,
            "performed_by": row.performed_by,
            "details": details,
            "created_at": str(row.created_at) if row.created_at else "",
        })
    emit({"ok": True, "data": data})


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("action", choices=["list", "save", "show", "export", "delete", "remit", "receive", "audit"])
    parser.add_argument("--payload", default="{}")
    args = parser.parse_args()

    try:
        payload = json.loads(args.payload or "{}")
        if args.action == "list":
            list_batches()
        if args.action == "save":
            save_batch(payload)
        if args.action == "show":
            get_batch(payload.get("report_no") or "")
        if args.action == "export":
            export_batch(payload.get("report_no") or "")
        if args.action == "delete":
            delete_batch(payload.get("report_no") or "")
        if args.action == "remit":
            remit_batch(payload)
        if args.action == "receive":
            receive_remittance(payload)
        if args.action == "audit":
            audit_trail(payload.get("report_no") or "")
    except Exception as exc:
        emit({"ok": False, "error": str(exc), "type": exc.__class__.__name__}, 1)


if __name__ == "__main__":
    main()
