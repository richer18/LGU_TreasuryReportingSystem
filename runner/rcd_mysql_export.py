import argparse
import json
import shutil
import sys
import tempfile
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font


PROJECT_ROOT = Path(__file__).resolve().parents[1]
TEMPLATE_PATH = PROJECT_ROOT / "template" / "RCD_UPDATED.xlsx"


def emit(payload, code=0):
    print(json.dumps(payload, ensure_ascii=False))
    raise SystemExit(code)


def safe_filename(value):
    cleaned = "".join(ch if ch.isalnum() or ch in "-_" else "_" for ch in str(value or "RCD"))
    while "__" in cleaned:
        cleaned = cleaned.replace("__", "_")
    return cleaned.strip("_") or "RCD"


def money(value):
    try:
        return round(float(str(value or 0).replace(",", "")), 2)
    except ValueError:
        return 0.0


def count_range(start, end):
    digits_start = "".join(ch for ch in str(start or "") if ch.isdigit())
    digits_end = "".join(ch for ch in str(end or start or "") if ch.isdigit())
    if not digits_start or not digits_end:
        return 0
    first = int(digits_start)
    last = int(digits_end)
    return last - first + 1 if last >= first else 0


def collector_name(value):
    mapping = {
        "FLORA MY": "FLORA MY D. FERRER",
        "AGNES": "AGNES B. ELLO",
        "RICARDO": "RICARDO T. ENOPIA",
        "IRIS": "ANGELIQUE IRIS A. RAFALES",
        "EMILY": "EMILY E. CREDO",
    }
    key = str(value or "").strip().upper()
    return mapping.get(key, value or "RCD")


def write_basic_workbook(path, batch):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "RCD"

    rows = [
        ["REPORT OF COLLECTIONS AND DEPOSIT"],
        ["Municipality of Zamboanguita"],
        [],
        ["RCD No.", batch.get("id") or "-"],
        ["Date", batch.get("date") or ""],
        ["Collector", collector_name(batch.get("collector"))],
        ["Fund", batch.get("fund") or ""],
        [],
        ["A. COLLECTIONS"],
        ["Type / Form No.", "OR From", "OR To", "OR Count", "Amount", "Validation"],
    ]

    for line in batch.get("lines") or []:
        rows.append([
            line.get("formType") or "",
            line.get("receiptFrom") or "",
            line.get("receiptTo") or "",
            count_range(line.get("receiptFrom"), line.get("receiptTo")),
            money(line.get("collectorAmount")),
            line.get("validationStatus") or "",
        ])

    rows.append(["", "", "", "TOTAL", money(batch.get("total")), ""])
    rows.append([])
    rows.append(["C. ACCOUNTABILITY OF ACCOUNTABLE FORMS"])
    rows.append(["Form", "Beginning Balance", "Receipt", "Issued", "Ending Balance"])

    for line in batch.get("lines") or []:
        rows.append([
            line.get("formType") or "",
            f"{line.get('beginningQty') or 0}: {line.get('beginningFrom') or ''} - {line.get('beginningTo') or ''}",
            f"{line.get('receiptAccountQty') or 0}: {line.get('receiptAccountFrom') or ''} - {line.get('receiptAccountTo') or ''}",
            f"{count_range(line.get('receiptFrom'), line.get('receiptTo'))}: {line.get('receiptFrom') or ''} - {line.get('receiptTo') or ''}",
            f"{line.get('endingQty') or ''}: {line.get('endingFrom') or ''} - {line.get('endingTo') or ''}",
        ])

    for row in rows:
        sheet.append(row)

    for cell in sheet[1]:
        cell.font = Font(bold=True, size=14)
        cell.alignment = Alignment(horizontal="center")
    sheet.merge_cells("A1:F1")
    sheet.merge_cells("A2:F2")
    sheet["A9"].font = Font(bold=True)
    sheet["A10"].font = Font(bold=True)
    for cell in sheet[10]:
        cell.font = Font(bold=True)
    total_row = 11 + len(batch.get("lines") or [])
    for cell in sheet[total_row]:
        cell.font = Font(bold=True)
    accountability_header = total_row + 3
    sheet[f"A{accountability_header}"].font = Font(bold=True)
    for cell in sheet[accountability_header + 1]:
        cell.font = Font(bold=True)
    for column in ["A", "B", "C", "D", "E", "F"]:
        sheet.column_dimensions[column].width = 22
    for row in sheet.iter_rows(min_col=5, max_col=5):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'

    workbook.save(path)


def fill_template(path, batch):
    workbook = load_workbook(path)
    if "E_GOV_ONLINE_PAYMENT" in workbook.sheetnames:
        workbook.remove(workbook["E_GOV_ONLINE_PAYMENT"])
    sheet = workbook["100_GF"] if "100_GF" in workbook.sheetnames else workbook.active

    sheet["A1"] = "REPORT OF COLLECTIONS AND DEPOSIT"
    sheet["A3"] = f"RCD No.: {batch.get('id') or '-'}"
    sheet["A4"] = f"Date: {batch.get('date') or ''}"
    sheet["A5"] = f"Collector: {collector_name(batch.get('collector'))}"

    start_row = 10
    sheet.cell(start_row, 1, "Type / Form No.")
    sheet.cell(start_row, 2, "OR From")
    sheet.cell(start_row, 3, "OR To")
    sheet.cell(start_row, 4, "Amount")
    for col in range(1, 5):
        sheet.cell(start_row, col).font = Font(bold=True)

    row = start_row + 1
    for line in batch.get("lines") or []:
        sheet.cell(row, 1, line.get("formType") or "")
        sheet.cell(row, 2, line.get("receiptFrom") or "")
        sheet.cell(row, 3, line.get("receiptTo") or "")
        sheet.cell(row, 4, money(line.get("collectorAmount")))
        sheet.cell(row, 4).number_format = '#,##0.00'
        row += 1

    sheet.cell(row, 3, "TOTAL")
    sheet.cell(row, 4, money(batch.get("total")))
    sheet.cell(row, 3).font = Font(bold=True)
    sheet.cell(row, 4).font = Font(bold=True)
    sheet.cell(row, 4).number_format = '#,##0.00'
    workbook.save(path)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--payload-file", required=True)
    args = parser.parse_args()

    with open(args.payload_file, "r", encoding="utf-8") as handle:
        batch = json.load(handle)

    out_dir = Path(tempfile.gettempdir()) / "lgu_treasury_rcd_mysql"
    out_dir.mkdir(parents=True, exist_ok=True)
    filename = f"{safe_filename(collector_name(batch.get('collector')) + '_' + str(batch.get('date') or ''))}.xlsx"
    path = out_dir / filename

    if TEMPLATE_PATH.exists():
        shutil.copy2(TEMPLATE_PATH, path)
        try:
            fill_template(path, batch)
        except Exception:
            write_basic_workbook(path, batch)
    else:
        write_basic_workbook(path, batch)

    emit({"ok": True, "path": str(path), "filename": filename})


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        emit({"ok": False, "error": str(exc), "type": exc.__class__.__name__}, 1)
