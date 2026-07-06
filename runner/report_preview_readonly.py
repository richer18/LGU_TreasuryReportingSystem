import argparse
import json
import os
import sys
from datetime import date, datetime, timedelta
from decimal import Decimal
from pathlib import Path

from firebird_probe import connect

APPDATA_SITE = os.environ.get("APPDATA")
USER_PROFILE = os.environ.get("USERPROFILE") or r"C:\Users\LIFT-LAPTOP"
USER_SITE_CANDIDATES = [
    Path(APPDATA_SITE) / "Python" / "Python314" / "site-packages" if APPDATA_SITE else None,
    Path(USER_PROFILE) / "AppData" / "Roaming" / "Python" / "Python314" / "site-packages",
]
for user_site in USER_SITE_CANDIDATES:
    if user_site and user_site.exists() and str(user_site) not in sys.path:
        sys.path.append(str(user_site))

BUSINESS_PERMIT_DIR = Path(__file__).resolve().parents[1] / "BUSINESS_PERMIT_REPORT"


SUMMARY_COLUMNS = [
    "source",
    "total_collections",
    "national",
    "provincial_general_fund",
    "provincial_sef",
    "provincial_total",
    "municipal_general_fund",
    "municipal_sef",
    "municipal_trust_fund",
    "municipal_total",
    "barangay_share",
    "fisheries",
]

NO_RPT_ORDER = [
    "Manufacturing",
    "Distributor",
    "Retailing",
    "Banks & Other Financial Int.",
    "Other Business Tax",
    "Sand & Gravel",
    "Fines & Penalties",
    "Mayor's Permit",
    "Weights & Measures",
    "Tricycle Permit Fee",
    "Occupation Tax",
    "Cert. of Ownership",
    "Cert. of Transfer",
    "Cockpit Share",
    "Docking and Mooring Fee",
    "Sultadas",
    "Miscellaneous",
    "Registration of Birth",
    "Marriage Fees",
    "Burial Fees",
    "Correction of Entry",
    "Fishing Permit Fee",
    "Sale of Agri. Prod.",
    "Sale of Acc. Forms",
    "Water Fees",
    "Market Stall Fee",
    "Cash Tickets",
    "Slaughterhouse Fee",
    "Rent of Equipment",
    "Doc Stamp Tax",
    "Secretary Fees",
    "Med./Lab. Fees",
    "Garbage Fees",
    "Cutting Tree",
    "Community Tax",
    "Building Permit Fee",
    "Electrical Permit Fee",
    "Zoning Fee",
    "Livestock",
    "Diving Fee",
]

PAID_PAYMENT_SQL = """
          AND COALESCE(p.VOID_BV, 0) = 0
          AND COALESCE(TRIM(p.STATUS_CT), '') NOT IN ('CNL', 'CAN', 'CNC', 'CANCEL', 'CANCELLED', 'VOID', 'VOI')
"""

BPLS_TAX_ON_BUSINESS_SOURCES = {
    "Manufacturing",
    "Distributor",
    "Retailing",
    "Banks & Other Financial Int.",
    "Other Business Tax",
    "Fines & Penalties",
}


def scalar(value):
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, str):
        return value.strip()
    return value


def normalize_source_name(name):
    aliases = {
        "Marriage Fees": "Marriage Fee",
        "Burial Fees": "Burial Fee",
        "Sale of Acc. Forms": "Sale of Acct. Forms",
        "Water Fees": "Water Fee",
        "Slaughterhouse Fee": "SlaughterHouse Fee",
        "Rent of Equipment": "Rental of Equipment",
        "Secretary Fees": "Secretaries Fees",
        "Community Tax": "Com Tax Cert.",
    }
    return aliases.get(name, name)


def classify_summary_source(itaxtype, source_id, source_ct):
    code = (itaxtype or "").strip()
    source_ct = (source_ct or "").strip()
    try:
        source_id = int(source_id) if source_id is not None else None
    except (TypeError, ValueError):
        source_id = None

    if source_ct in ("CTCI", "CTCC") or code == "CTC":
        return "Community Tax"
    if code == "MAS":
        return "Manufacturing"
    if code == "WHO":
        return "Distributor"
    if code == "RET":
        return "Retailing"
    if code == "BFI":
        return "Banks & Other Financial Int."
    if code in ("OBT", "CIC", "PED", "EMD") and source_id not in (807, 808):
        return "Other Business Tax"
    if code in ("TSG", "TSB"):
        return "Sand & Gravel"
    if code == "FPT":
        return "Fines & Penalties"
    if code == "MP":
        return "Mayor's Permit"
    if code == "FWM":
        return "Weights & Measures"
    if code in ("TOP", "MTO", "FLF"):
        return "Tricycle Permit Fee"
    if code == "OCC":
        return "Occupation Tax"
    if code == "COO":
        return "Cert. of Ownership"
    if code == "COT":
        return "Cert. of Transfer"
    if code == "CS":
        return "Cockpit Share"
    if code == "FRF" and source_id in (580, 639):
        return "Docking and Mooring Fee"
    if code in ("ST", "ATM"):
        return "Sultadas"
    if code in ("IM", "OPF", "SBF") or source_id in (807, 808):
        return "Miscellaneous"
    if code == "RB":
        return "Registration of Birth"
    if code == "RM":
        return "Marriage Fees"
    if code == "BF":
        return "Burial Fees"
    if code == "CE":
        return "Correction of Entry"
    if code in ("FRF", "IF"):
        return "Fishing Permit Fee"
    if code == "IAP":
        return "Sale of Agri. Prod."
    if code == "IAF":
        return "Sale of Acc. Forms"
    if code in ("WTR", "IWO"):
        return "Water Fees"
    if code in ("RFM", "MSF"):
        return "Market Stall Fee"
    if code in ("SPF", "RFS"):
        return "Slaughterhouse Fee"
    if code in ("RFR", "IPG", "ICO"):
        return "Rent of Equipment"
    if code == "SF" and source_id == 810:
        return "Doc Stamp Tax"
    if code in ("SF", "PCL", "HEC", "OCL"):
        return "Secretary Fees"
    if code == "MDL":
        return "Med./Lab. Fees"
    if code == "GCF":
        return "Garbage Fees"
    if code in ("PFB", "BUF", "INS"):
        return "Building Permit Fee"
    if code == "EP":
        return "Electrical Permit Fee"
    if code == "ZLC":
        return "Zoning Fee"
    if code == "IFL":
        return "Livestock"
    if code == "IFD":
        return "Diving Fee"
    return None


def empty_summary_row(source):
    return {
        "source": normalize_source_name(source),
        "total_collections": Decimal("0"),
        "national": Decimal("0"),
        "provincial_general_fund": Decimal("0"),
        "provincial_sef": Decimal("0"),
        "provincial_total": Decimal("0"),
        "municipal_general_fund": Decimal("0"),
        "municipal_sef": Decimal("0"),
        "municipal_trust_fund": Decimal("0"),
        "municipal_total": Decimal("0"),
        "barangay_share": Decimal("0"),
        "fisheries": Decimal("0"),
    }


def split_summary_amount(name, amount):
    amount = amount or Decimal("0")
    row = empty_summary_row(name)

    if name == "Cockpit Share":
        row["total_collections"] = amount
        row["provincial_general_fund"] = amount * Decimal("0.50")
        row["provincial_total"] = row["provincial_general_fund"]
        row["municipal_general_fund"] = amount * Decimal("0.50")
        row["municipal_total"] = row["municipal_general_fund"]
    elif name == "Building Permit Fee":
        row["total_collections"] = amount
        row["national"] = amount * Decimal("0.05")
        row["municipal_general_fund"] = amount * Decimal("0.80")
        row["municipal_trust_fund"] = amount * Decimal("0.15")
        row["municipal_total"] = row["municipal_general_fund"] + row["municipal_trust_fund"]
    elif name == "Livestock":
        row["total_collections"] = amount
        row["national"] = amount * Decimal("0.20")
        row["municipal_general_fund"] = amount * Decimal("0.80")
        row["municipal_total"] = row["municipal_general_fund"]
    elif name == "Diving Fee":
        row["total_collections"] = amount
        row["municipal_general_fund"] = amount * Decimal("0.40")
        row["municipal_total"] = row["municipal_general_fund"]
        row["barangay_share"] = amount * Decimal("0.30")
        row["fisheries"] = amount * Decimal("0.30")
    else:
        row["total_collections"] = amount
        row["municipal_general_fund"] = amount
        row["municipal_total"] = amount

    return row


def fetch_no_rpt_summary(date_from, date_to):
    amounts = {name: Decimal("0") for name in NO_RPT_ORDER}
    sql = f"""
        SELECT
            pd.ITAXTYPE_CT,
            pd.SOURCEID,
            pd.SOURCE_CT,
            SUM(pd.AMOUNTPAID) AS AMOUNT
        FROM PAYMENT p
        JOIN PAYMENTDETAIL pd ON pd.PAYMENT_ID = p.PAYMENT_ID
        WHERE p.PAYMENTDATE >= CAST(? AS DATE)
          AND p.PAYMENTDATE < DATEADD(1 DAY TO CAST(? AS DATE))
          {PAID_PAYMENT_SQL}
          AND COALESCE(p.PAYGROUP_CT, '') <> 'RPT'
        GROUP BY pd.ITAXTYPE_CT, pd.SOURCEID, pd.SOURCE_CT
    """

    connection = connect()
    try:
        cursor = connection.cursor()
        cursor.execute(sql, (date_from, date_to))
        for itaxtype, source_id, source_ct, amount in cursor.fetchall():
            source = classify_summary_source(itaxtype, source_id, source_ct)
            if source:
                amounts[source] = amounts.get(source, Decimal("0")) + (amount or Decimal("0"))
        connection.rollback()
    finally:
        connection.close()

    for row in fetch_tax_on_business_summary(date_from, date_to):
        category = row.get("category")
        if category in BPLS_TAX_ON_BUSINESS_SOURCES:
            amounts[category] = row.get("total", Decimal("0")) or Decimal("0")

    return [split_summary_amount(name, amounts.get(name, Decimal("0"))) for name in NO_RPT_ORDER]


def fetch_rpt_buckets(date_from, date_to):
    sql = f"""
        SELECT
            pcd.PROPERTYKIND_CT,
            pcd.ITAXTYPE_CT,
            pcd.CASETYPE_CT,
            pcd.TAXYEAR,
            SUM(pcd.AMOUNT) AS AMOUNT
        FROM PAYMENT p
        JOIN PAYMENTCLASSDETAIL pcd ON pcd.PAYMENT_ID = p.PAYMENT_ID
        WHERE p.PAYMENTDATE >= CAST(? AS DATE)
          AND p.PAYMENTDATE < DATEADD(1 DAY TO CAST(? AS DATE))
          AND p.PAYGROUP_CT = 'RPT'
          {PAID_PAYMENT_SQL}
          AND COALESCE(pcd.CANCELLED_BV, 0) = 0
        GROUP BY pcd.PROPERTYKIND_CT, pcd.ITAXTYPE_CT, pcd.CASETYPE_CT, pcd.TAXYEAR
    """

    entries = []
    connection = connect()
    try:
        cursor = connection.cursor()
        cursor.execute(sql, (date_from, date_to))
        for propkind, itaxtype, casetype, taxyear, amount in cursor.fetchall():
            entries.append({
                "property_group": "Land" if (propkind or "").strip() == "L" else "Bldg.",
                "tax_type": (itaxtype or "").strip(),
                "case_type": (casetype or "").strip(),
                "taxyear": taxyear,
                "amount": amount or Decimal("0"),
            })
        connection.rollback()
    finally:
        connection.close()

    report_year = datetime.strptime(date_from, "%Y-%m-%d").year
    buckets = {}

    for entry in entries:
        taxyear = entry["taxyear"]
        if taxyear and taxyear > report_year:
            continue

        if entry["case_type"] == "PEN":
            line = "Penalties"
        elif taxyear == report_year:
            line = "Current Year"
        else:
            line = "Previous Years"

        key = (entry["property_group"], entry["tax_type"], line)
        buckets[key] = buckets.get(key, Decimal("0")) + entry["amount"]

    return buckets


def rpt_summary_row(label, amount, tax_type):
    row = empty_summary_row(label)
    amount = amount or Decimal("0")
    row["total_collections"] = amount

    if tax_type == "BSC":
        row["provincial_general_fund"] = amount * Decimal("0.35")
        row["provincial_total"] = row["provincial_general_fund"]
        row["municipal_general_fund"] = amount * Decimal("0.40")
        row["municipal_total"] = row["municipal_general_fund"]
        row["barangay_share"] = amount * Decimal("0.25")
    else:
        row["provincial_sef"] = amount * Decimal("0.50")
        row["provincial_total"] = row["provincial_sef"]
        row["municipal_sef"] = amount * Decimal("0.50")
        row["municipal_total"] = row["municipal_sef"]

    return row


def fetch_rpt_summary(date_from, date_to):
    buckets = fetch_rpt_buckets(date_from, date_to)
    layout = [
        ("Real Property Tax - Basic/Land", None, None),
        ("Current Year", "Land", "BSC"),
        ("Previous Years", "Land", "BSC"),
        ("Penalties", "Land", "BSC"),
        ("Real Property Tax - SEF/Land", None, None),
        ("Current Year", "Land", "SEF"),
        ("Previous Years", "Land", "SEF"),
        ("Penalties", "Land", "SEF"),
        ("Real Property Tax - Basic/Bldg.", None, None),
        ("Current Year", "Bldg.", "BSC"),
        ("Previous Years", "Bldg.", "BSC"),
        ("Penalties", "Bldg.", "BSC"),
        ("Real Property Tax - SEF/Bldg.", None, None),
        ("Current Year", "Bldg.", "SEF"),
        ("Previous Years", "Bldg.", "SEF"),
        ("Penalties", "Bldg.", "SEF"),
    ]
    rows = []

    for label, group, tax_type in layout:
        if group is None:
            rows.append({"source": label, "section": True})
        else:
            rows.append(rpt_summary_row(label, buckets.get((group, tax_type, label), Decimal("0")), tax_type))

    return rows


def fetch_rpt_sharing_summary(date_from, date_to):
    sql = f"""
        SELECT
            pcd.PROPERTYKIND_CT,
            pcd.CASETYPE_CT,
            pcd.TAXYEAR,
            SUM(pcd.AMOUNT) AS AMOUNT
        FROM PAYMENT p
        JOIN PAYMENTCLASSDETAIL pcd ON pcd.PAYMENT_ID = p.PAYMENT_ID
        WHERE p.PAYMENTDATE >= CAST(? AS DATE)
          AND p.PAYMENTDATE < DATEADD(1 DAY TO CAST(? AS DATE))
          AND p.PAYGROUP_CT = 'RPT'
          {PAID_PAYMENT_SQL}
          AND COALESCE(pcd.CANCELLED_BV, 0) = 0
          AND pcd.ITAXTYPE_CT = 'BSC'
        GROUP BY pcd.PROPERTYKIND_CT, pcd.CASETYPE_CT, pcd.TAXYEAR
    """

    entries = []
    connection = connect()
    try:
        cursor = connection.cursor()
        cursor.execute(sql, (date_from, date_to))
        for property_kind, case_type, taxyear, amount in cursor.fetchall():
            entries.append({
                "property_group": "Land" if (property_kind or "").strip() == "L" else "Building",
                "case_type": (case_type or "").strip(),
                "taxyear": taxyear,
                "amount": amount or Decimal("0"),
            })
        connection.rollback()
    finally:
        connection.close()

    report_year = datetime.strptime(date_from, "%Y-%m-%d").year
    buckets = {
        "Land": {"Current": Decimal("0"), "Prior": Decimal("0"), "Penalties": Decimal("0")},
        "Building": {"Current": Decimal("0"), "Prior": Decimal("0"), "Penalties": Decimal("0")},
    }

    for entry in entries:
        group = entry["property_group"]
        amount = entry["amount"]
        taxyear = entry["taxyear"]

        if taxyear and taxyear > report_year:
            continue

        if entry["case_type"] == "PEN":
            buckets[group]["Penalties"] += amount
        elif entry["case_type"] == "DED":
            if taxyear == report_year:
                buckets[group]["Current"] -= abs(amount)
            else:
                buckets[group]["Prior"] -= abs(amount)
        elif taxyear == report_year:
            buckets[group]["Current"] += amount
        else:
            buckets[group]["Prior"] += amount

    rows = []
    grand_total = Decimal("0")
    for group in ("Land", "Building"):
        group_total = Decimal("0")
        for category in ("Current", "Prior", "Penalties"):
            amount = buckets[group][category]
            group_total += amount
            rows.append({
                "property_group": group,
                "category": category,
                "bsc_amount": amount,
                "provincial_share_35": amount * Decimal("0.35"),
                "municipal_share_40": amount * Decimal("0.40"),
                "barangay_share_25": amount * Decimal("0.25"),
            })

        grand_total += group_total
        rows.append({
            "property_group": group,
            "category": "TOTAL",
            "bsc_amount": group_total,
            "provincial_share_35": group_total * Decimal("0.35"),
            "municipal_share_40": group_total * Decimal("0.40"),
            "barangay_share_25": group_total * Decimal("0.25"),
            "total": True,
        })

    rows.append({
        "property_group": "Land and Building",
        "category": "GRAND TOTAL",
        "bsc_amount": grand_total,
        "provincial_share_35": grand_total * Decimal("0.35"),
        "municipal_share_40": grand_total * Decimal("0.40"),
        "barangay_share_25": grand_total * Decimal("0.25"),
        "grand_total": True,
    })
    return rows


def sharing_row_for_classification(property_kind, class_code):
    property_kind = (property_kind or "").strip()
    class_code = (class_code or "").strip()
    if property_kind == "L":
        if class_code == "A":
            return 11
        if class_code == "R":
            return 12
        if class_code == "C":
            return 13
        return 14

    if property_kind == "M":
        return 22
    if class_code == "R":
        return 23
    if class_code == "C":
        return 24
    if class_code == "A":
        return 25
    if class_code.upper().startswith("S"):
        return 26
    return 26


def fetch_summary_sharing_template_cells(date_from, date_to):
    sql = f"""
        SELECT
            pcd.PROPERTYKIND_CT,
            COALESCE(pcd.CLASSCODE_CT, ra.PREDOMCLASSCODE_CT) AS CLASSCODE_CT,
            pcd.ITAXTYPE_CT,
            pcd.CASETYPE_CT,
            pcd.TAXYEAR,
            SUM(pcd.AMOUNT) AS AMOUNT
        FROM PAYMENT p
        JOIN PAYMENTCLASSDETAIL pcd ON pcd.PAYMENT_ID = p.PAYMENT_ID
        LEFT JOIN RPTASSESSMENT ra ON ra.TAXTRANS_ID = pcd.TAXTRANS_ID
        WHERE p.PAYMENTDATE >= CAST(? AS DATE)
          AND p.PAYMENTDATE < DATEADD(1 DAY TO CAST(? AS DATE))
          AND p.PAYGROUP_CT = 'RPT'
          {PAID_PAYMENT_SQL}
          AND COALESCE(pcd.CANCELLED_BV, 0) = 0
        GROUP BY pcd.PROPERTYKIND_CT, COALESCE(pcd.CLASSCODE_CT, ra.PREDOMCLASSCODE_CT),
                 pcd.ITAXTYPE_CT, pcd.CASETYPE_CT, pcd.TAXYEAR
    """
    values = {}
    report_year = datetime.strptime(date_from, "%Y-%m-%d").year

    connection = connect()
    try:
        cursor = connection.cursor()
        cursor.execute(sql, (date_from, date_to))
        fetched_rows = cursor.fetchall()

        for property_kind, class_code, tax_type, case_type, taxyear, amount in fetched_rows:
            if taxyear and taxyear > report_year:
                continue

            row_index = sharing_row_for_classification(property_kind, class_code)
            tax_type = (tax_type or "").strip()
            case_type = (case_type or "").strip()
            if tax_type == "BSC":
                current_col, discount_col, prior_col, pen_current_col, pen_prior_col = 3, 4, 5, 6, 7
            elif tax_type == "SEF":
                current_col, discount_col, prior_col, pen_current_col, pen_prior_col = 10, 11, 12, 13, 14
            else:
                continue

            if case_type == "DED":
                col_index = discount_col
                value = abs(amount or Decimal("0"))
            elif case_type == "PEN":
                col_index = pen_current_col if taxyear == report_year else pen_prior_col
                value = amount or Decimal("0")
            else:
                col_index = current_col if taxyear == report_year else prior_col
                value = amount or Decimal("0")
            values[(row_index, col_index)] = values.get((row_index, col_index), Decimal("0")) + value
        connection.rollback()
    finally:
        connection.close()

    cells = []
    for row_index in (11, 12, 13, 14, 22, 23, 24, 25, 26):
        for col_index in (3, 4, 5, 6, 7, 10, 11, 12, 13, 14):
            cells.append({
                "row": row_index,
                "column": col_index,
                "value": values.get((row_index, col_index), Decimal("0")),
            })
    return cells


def add_full_daily_amount(daily, day, column_name, amount):
    if hasattr(day, "date"):
        day = day.date()
    if day not in daily:
        daily[day] = {
            "ctc": Decimal("0"),
            "rpt": Decimal("0"),
            "gf_tf": Decimal("0"),
        }
    daily[day][column_name] += amount or Decimal("0")


def fetch_full_report_collections(date_from, date_to):
    daily = {}
    ctc_sql = f"""
        SELECT
            CAST(p.PAYMENTDATE AS DATE) AS COLLECTION_DATE,
            SUM(pd.AMOUNTPAID) AS AMOUNT
        FROM PAYMENT p
        JOIN PAYMENTDETAIL pd ON pd.PAYMENT_ID = p.PAYMENT_ID
        WHERE p.PAYMENTDATE >= CAST(? AS DATE)
          AND p.PAYMENTDATE < DATEADD(1 DAY TO CAST(? AS DATE))
          {PAID_PAYMENT_SQL}
          AND (pd.SOURCE_CT IN ('CTCI', 'CTCC') OR pd.ITAXTYPE_CT = 'CTC')
        GROUP BY CAST(p.PAYMENTDATE AS DATE)
    """
    rpt_sql = f"""
        SELECT
            CAST(p.PAYMENTDATE AS DATE) AS COLLECTION_DATE,
            SUM(pcd.AMOUNT) AS AMOUNT
        FROM PAYMENT p
        JOIN PAYMENTCLASSDETAIL pcd ON pcd.PAYMENT_ID = p.PAYMENT_ID
        WHERE p.PAYMENTDATE >= CAST(? AS DATE)
          AND p.PAYMENTDATE < DATEADD(1 DAY TO CAST(? AS DATE))
          AND p.PAYGROUP_CT = 'RPT'
          {PAID_PAYMENT_SQL}
          AND COALESCE(pcd.CANCELLED_BV, 0) = 0
        GROUP BY CAST(p.PAYMENTDATE AS DATE)
    """
    gf_tf_sql = f"""
        SELECT
            CAST(p.PAYMENTDATE AS DATE) AS COLLECTION_DATE,
            SUM(pd.AMOUNTPAID) AS AMOUNT
        FROM PAYMENT p
        JOIN PAYMENTDETAIL pd ON pd.PAYMENT_ID = p.PAYMENT_ID
        WHERE p.PAYMENTDATE >= CAST(? AS DATE)
          AND p.PAYMENTDATE < DATEADD(1 DAY TO CAST(? AS DATE))
          {PAID_PAYMENT_SQL}
          AND COALESCE(p.PAYGROUP_CT, '') <> 'RPT'
          AND NOT (pd.SOURCE_CT IN ('CTCI', 'CTCC') OR pd.ITAXTYPE_CT = 'CTC')
        GROUP BY CAST(p.PAYMENTDATE AS DATE)
    """

    connection = connect()
    try:
        cursor = connection.cursor()
        for sql, column_name in ((ctc_sql, "ctc"), (rpt_sql, "rpt"), (gf_tf_sql, "gf_tf")):
            cursor.execute(sql, (date_from, date_to))
            for collection_date, amount in cursor.fetchall():
                add_full_daily_amount(daily, collection_date, column_name, amount)
        connection.rollback()
    finally:
        connection.close()

    start_date = datetime.strptime(date_from, "%Y-%m-%d").date()
    end_date = datetime.strptime(date_to, "%Y-%m-%d").date()
    rows = []
    current = start_date
    while current <= end_date:
        amounts = daily.get(current, {"ctc": Decimal("0"), "rpt": Decimal("0"), "gf_tf": Decimal("0")})
        total = amounts["ctc"] + amounts["rpt"] + amounts["gf_tf"]
        rows.append({
            "date": current,
            "ctc": amounts["ctc"],
            "rpt": amounts["rpt"],
            "gf_tf": amounts["gf_tf"],
            "due_from": "",
            "rcd_total": total,
        })
        current += timedelta(days=1)
    return rows


def clean_text(value):
    if value is None:
        return ""
    return str(value).strip()


def decimal_value(value):
    if value in (None, ""):
        return Decimal("0")
    try:
        return Decimal(str(value).replace(",", ""))
    except Exception:
        return Decimal("0")


def find_business_permit_workbook(pattern):
    matches = sorted(BUSINESS_PERMIT_DIR.glob(pattern))
    if not matches:
        raise FileNotFoundError(f"Business permit workbook was not found: {BUSINESS_PERMIT_DIR / pattern}")
    return matches[-1]


def parse_excel_date(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        from openpyxl.utils.datetime import from_excel
        return from_excel(value).date()
    text = clean_text(value)
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%B %d, %Y", "%d-%b-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def load_sheet_records(path):
    from openpyxl import load_workbook
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    headers = [clean_text(value) for value in rows[0]]
    records = []
    for row in rows[1:]:
        records.append({headers[index]: value for index, value in enumerate(row) if index < len(headers)})
    workbook.close()
    return records


def load_sheet_records_with_header(path, header_row):
    from openpyxl import load_workbook
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    headers = [clean_text(cell.value) for cell in next(sheet.iter_rows(min_row=header_row, max_row=header_row))]
    records = []
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        records.append({headers[index]: value for index, value in enumerate(row) if index < len(headers)})
    workbook.close()
    return records


def business_establishment_match_lookup(records):
    by_or = {}
    by_business_id = {}
    for record in records:
        or_number = clean_text(record.get("OR Number"))
        business_id = clean_text(record.get("Business Identification Number"))
        has_paid = decimal_value(record.get("Total Amount Paid")) > 0
        has_permit = bool(clean_text(record.get("Permit No.")))
        score = (1 if has_paid else 0, 1 if has_permit else 0)

        if or_number:
            existing_score, _existing = by_or.get(or_number, ((-1, -1), None))
            if score > existing_score:
                by_or[or_number] = (score, record)
        if business_id:
            existing_score, _existing = by_business_id.get(business_id, ((-1, -1), None))
            if score > existing_score:
                by_business_id[business_id] = (score, record)

    return (
        {or_number: record for or_number, (_score, record) in by_or.items()},
        {business_id: record for business_id, (_score, record) in by_business_id.items()},
    )


def tax_on_business_category(business_nature, business_line):
    nature_text = clean_text(business_nature).lower()
    line_text = clean_text(business_line).lower()
    text = f"{nature_text} {line_text}"
    if any(keyword in text for keyword in ("bank", "financial", "lending", "pawn", "money", "remittance", "insurance")):
        return "Banks & Other Financial Int."
    if any(keyword in text for keyword in ("manufactur", "baking", "bakery", "milling", "printing", "processed")):
        return "Manufacturing"
    if any(keyword in line_text for keyword in ("wholesale", "distributor", "distribution")):
        return "Distributor"
    if any(keyword in line_text for keyword in ("retail", "store", "sari-sari", "pharmacy", "hardware", "convenience")):
        return "Retailing"
    if "wholesale and retail trade" in nature_text:
        return "Retailing"
    return "Other Business Tax"


def fetch_tax_on_business_summary(date_from, date_to):
    abstract_path = find_business_permit_workbook("ABSTRACT_OF_GENERAL_COLLECTION-BPLS*.xlsx")
    establishment_path = find_business_permit_workbook("BUSINESS_ESTABLISHMENT-BPLS*.xlsx")
    abstract_records = load_sheet_records_with_header(abstract_path, 7)
    establishment_records = load_sheet_records(establishment_path)
    by_or, by_business_id = business_establishment_match_lookup(establishment_records)
    start_date = datetime.strptime(date_from, "%Y-%m-%d").date()
    end_date = datetime.strptime(date_to, "%Y-%m-%d").date()
    category_order = [
        "Manufacturing",
        "Distributor",
        "Retailing",
        "Banks & Other Financial Int.",
        "Other Business Tax",
        "Fines & Penalties",
    ]
    summary = {
        category: {"business_tax": Decimal("0"), "surcharge": Decimal("0")}
        for category in category_order
    }

    for record in abstract_records:
        or_date = parse_excel_date(record.get("O.R. Date"))
        if or_date is None or not (start_date <= or_date <= end_date):
            continue
        business_tax = decimal_value(record.get("Business Tax"))
        surcharge = decimal_value(record.get("Surcharge"))
        if business_tax == 0 and surcharge == 0:
            continue

        or_number = clean_text(record.get("O.R. Number"))
        business_id = clean_text(record.get("Business Identification Number"))
        establishment = by_or.get(or_number) or by_business_id.get(business_id) or {}
        category = tax_on_business_category(
            establishment.get("Business Nature"),
            establishment.get("Business Line"),
        )

        if business_tax:
            summary[category]["business_tax"] += business_tax
        if surcharge:
            summary["Fines & Penalties"]["surcharge"] += surcharge

    rows = []
    for category in category_order:
        business_tax = summary[category]["business_tax"]
        surcharge = summary[category]["surcharge"]
        rows.append({
            "category": category,
            "business_tax": business_tax,
            "surcharge": surcharge,
            "total": business_tax + surcharge,
        })
    return rows


def add_totals(rows):
    totals = {column: Decimal("0") for column in SUMMARY_COLUMNS if column != "source"}

    for row in rows:
        if row.get("section"):
            continue
        for column in totals:
            totals[column] += Decimal(str(row.get(column, 0) or 0))

    total_row = {"source": "TOTAL", **totals, "total": True}
    return rows + [total_row]


def build_report(number, date_from, date_to):
    if number == 21:
        rows = fetch_no_rpt_summary(date_from, date_to) + fetch_rpt_summary(date_from, date_to)
    elif number == 22:
        rows = fetch_no_rpt_summary(date_from, date_to)
    elif number == 23:
        rows = fetch_rpt_summary(date_from, date_to)
    elif number == 27:
        rows = fetch_rpt_sharing_summary(date_from, date_to)
        return {
            "ok": True,
            "mode": "read_only_report_preview",
            "report_number": number,
            "date_from": date_from,
            "date_to": date_to,
            "columns": [
                "property_group",
                "category",
                "bsc_amount",
                "provincial_share_35",
                "municipal_share_40",
                "barangay_share_25",
            ],
            "rows": rows,
            "template_cells": fetch_summary_sharing_template_cells(date_from, date_to),
        }
    elif number == 28:
        rows = fetch_rpt_sharing_summary(date_from, date_to)
        return {
            "ok": True,
            "mode": "read_only_report_preview",
            "report_number": number,
            "date_from": date_from,
            "date_to": date_to,
            "columns": [
                "property_group",
                "category",
                "bsc_amount",
                "provincial_share_35",
                "municipal_share_40",
                "barangay_share_25",
            ],
            "rows": rows,
            "template_cells": fetch_summary_sharing_template_cells(date_from, date_to),
        }
    elif number == 31:
        rows = fetch_full_report_collections(date_from, date_to)
        return {
            "ok": True,
            "mode": "read_only_report_preview",
            "report_number": number,
            "date_from": date_from,
            "date_to": date_to,
            "columns": ["date", "ctc", "rpt", "gf_tf", "due_from", "rcd_total"],
            "rows": rows,
        }
    elif number == 33:
        rows = fetch_tax_on_business_summary(date_from, date_to)
        return {
            "ok": True,
            "mode": "read_only_report_preview",
            "report_number": number,
            "date_from": date_from,
            "date_to": date_to,
            "columns": ["category", "business_tax", "surcharge", "total"],
            "rows": rows,
        }
    else:
        raise ValueError(f"Report {number} preview is not connected to the Firebird runner yet.")

    return {
        "ok": True,
        "mode": "read_only_report_preview",
        "report_number": number,
        "date_from": date_from,
        "date_to": date_to,
        "columns": SUMMARY_COLUMNS,
        "rows": add_totals(rows),
    }


def main():
    parser = argparse.ArgumentParser(description="Read-only Firebird report preview runner.")
    parser.add_argument("report_number", type=int)
    parser.add_argument("--date-from", required=True)
    parser.add_argument("--date-to", required=True)
    args = parser.parse_args()

    try:
        payload = build_report(args.report_number, args.date_from, args.date_to)
        print(json.dumps(payload, default=scalar))
        return 0
    except Exception as exc:
        print(json.dumps({
            "ok": False,
            "mode": "read_only_report_preview",
            "error": str(exc),
            "error_type": exc.__class__.__name__,
        }))
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
