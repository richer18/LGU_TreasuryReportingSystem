from pathlib import Path
from openpyxl import load_workbook
root = Path(r"\\MAIN-SERVER\LGU_TreasuryReportingSystem_$")
wb = load_workbook(root / "template" / "RCD_UPDATED.xlsx")
for sheet_name in ["100_GF", "200_SEF"]:
    ws = wb[sheet_name]
    print(f"--- {sheet_name} ---")
    for r in range(1, 45):
        vals = [ws.cell(r, c).value for c in range(1, 14)]
        if any(v not in (None, "") for v in vals):
            print(r, vals)
